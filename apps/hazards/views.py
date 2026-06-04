from urllib import request

from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import DeleteView, ListView, CreateView, UpdateView, DetailView, TemplateView
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Q, Value
from django.http import JsonResponse
from django.utils import timezone
from apps.organizations.models import *
from .models import Hazard, HazardPhoto, HazardVideo, HazardActionItem
from django.utils.safestring import mark_safe  # ADD THIS IMPORT

from django.contrib.auth import get_user_model
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .utils import generate_hazard_pdf
from django.views import View
from apps.common.image_utils import compress_image, compress_video

import json
from django.db.models import Count
from django.db.models.functions import Coalesce
from django.db.models.functions import TruncMonth
from .forms import HazardForm
from apps.notifications.services import NotificationService

# Make sure all models are imported
from apps.organizations.models import Plant, Zone, Location, SubLocation
from dateutil.relativedelta import relativedelta
import colorsys

User = get_user_model()


HAZARD_STATUS_FILTER_CHOICES = list(Hazard.STATUS_CHOICES) + [
    ('OVERDUE', 'Overdue'),
    ('CLOSED_LATE', 'Closed Late'),
]


def filter_hazards_by_status(queryset, selected_status):
    """Apply computed hazard status filters without changing the model schema."""
    if not selected_status:
        return queryset

    normalized_status = selected_status.upper()

    if normalized_status == 'OPEN':
        return queryset.exclude(status='CLOSED')
    if normalized_status in {'OVERDUE', 'CLOSED', 'CLOSED_LATE', 'LATE_CLOSED'}:
        target_status = 'CLOSED_LATE' if normalized_status == 'LATE_CLOSED' else normalized_status
        matched_hazard_ids = [
            hazard.pk for hazard in queryset.prefetch_related('action_items')
            if hazard.effective_status == target_status
        ]
        return queryset.filter(pk__in=matched_hazard_ids)

    return queryset.filter(status=selected_status)


class HazardDashboardView(LoginRequiredMixin, TemplateView):
    """Hazard Management Dashboard"""
    template_name = 'hazards/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        # Get hazards based on user role (This part is already correct)
        if user.is_superuser or (user.role and user.role.name == 'ADMIN'):
            hazards = Hazard.objects.all()
        elif user.get_all_plants():
            hazards = Hazard.objects.filter(plant__in=user.get_all_plants()).distinct()
        else:
            hazards = Hazard.objects.filter(reported_by=user)
        
        # Statistics (This part is already correct)
        context['total_hazards'] = hazards.count()
        context['open_hazards'] = hazards.exclude(status='CLOSED').count()
        context['this_month_hazards'] = hazards.filter(
            incident_datetime__month=datetime.date.today().month,
            incident_datetime__year=datetime.date.today().year
        ).count()
        
        # --- THIS IS THE SECTION TO UPDATE ---
        # Match the context variable names to your template (e.g., 'low_risk' instead of 'low_severity')
        context['critical_hazards'] = hazards.filter(severity='critical').count()
        context['low_risk'] = hazards.filter(severity='low').count()          # <-- UPDATED
        context['medium_risk'] = hazards.filter(severity='medium').count()    # <-- UPDATED
        context['high_risk'] = hazards.filter(severity='high').count()        # <-- UPDATED
        

        # Recent hazards (This part is already correct)
        context['recent_hazards'] = hazards.order_by('-incident_datetime')[:10]
        
        return context


class HazardListView(LoginRequiredMixin, ListView):
    """
    List all hazards with filtering.
    This view now includes specific logic to restrict data visibility based on user roles.
    - ADMIN/Superuser can see all hazards.
    - EMPLOYEE can only see hazards they have personally reported.
    - Other roles (like PLANT HEAD) see hazards related to their assigned plant.
    """
    model = Hazard
    template_name = 'hazards/hazard_list.html'
    context_object_name = 'hazards'
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user
        
        queryset = Hazard.objects.select_related('plant', 'location', 'reported_by').order_by('-incident_datetime')

        # Role-based filtering
        if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == 'ADMIN'):
            pass
        elif hasattr(user, 'role') and user.role and user.role.name == 'EMPLOYEE':
            queryset = queryset.filter(reported_by=user)
        elif user.get_all_plants():
            queryset = queryset.filter(plant__in=user.get_all_plants()).distinct()
        else:
            queryset = queryset.filter(reported_by=user)

        # Get filter parameters
        search = self.request.GET.get('search', '')
        hazard_type = self.request.GET.get('hazard_type', '')
        risk_level = self.request.GET.get('risk_level', '')
        status = self.request.GET.get('status', '')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        assigned_by = self.request.GET.get('assigned_by', '')
        assigned_to = self.request.GET.get('assigned_to', '')

        # Apply filters
        if search:
            queryset = queryset.filter(
                Q(report_number__icontains=search) |
                Q(hazard_title__icontains=search)
            )
        if hazard_type:
            queryset = queryset.filter(hazard_type=hazard_type)
        if risk_level:
            queryset = queryset.filter(severity=risk_level)
        queryset = filter_hazards_by_status(queryset, status)
        if assigned_by:
            queryset = queryset.filter(reported_by_id=assigned_by)

        if assigned_to:
            selected_user = User.objects.filter(id=assigned_to).first()
            if selected_user:
                queryset = queryset.filter(action_items__responsible_emails__icontains=selected_user.email).distinct()
        
        if date_from:
            queryset = queryset.filter(incident_datetime__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(incident_datetime__date__lte=date_to)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        reported_users = User.objects.filter(hazards_reported__isnull=False)

        assigned_users = User.objects.filter(hazards_assigned__isnull=False)

        context['assigned_by_users'] = User.objects.filter(
            hazards_reported__isnull=False,
            is_active=True,
            is_superuser=False,
            is_active_employee=True
        ).distinct().order_by('first_name', 'last_name')


        from apps.hazards.models import HazardActionItem
        selected_assigned_by = self.request.GET.get('assigned_by', '')
        selected_assigned_to = self.request.GET.get('assigned_to', '')
        action_items_queryset = HazardActionItem.objects.exclude(responsible_emails='')

        if selected_assigned_by:
            action_items_queryset = action_items_queryset.filter(
                hazard__reported_by_id=selected_assigned_by
            )

        assigned_emails = []

        for action in action_items_queryset:
            emails = [
                email.strip()
                for email in action.responsible_emails.split(',')
                if email.strip()
            ]
            assigned_emails.extend(emails)

        context['assigned_to_users'] = User.objects.filter(
            email__in=assigned_emails,
            is_active=True,
            is_superuser=False,
            is_active_employee=True
        ).distinct().order_by('first_name', 'last_name')

        if selected_assigned_to:
            valid_user_exists = context['assigned_to_users'].filter(
                id=selected_assigned_to
            ).exists()

            if not valid_user_exists:
                selected_assigned_to = ''

        context['assigned_to_users'] = User.objects.filter(
            email__in=assigned_emails,
            is_active=True,
            is_superuser=False,
            is_active_employee=True
        ).distinct().order_by('first_name', 'last_name')
        
        print("context['assigned_by_users']",context['assigned_by_users'])
        print("context['assigned_to_users']",context['assigned_to_users'])
        # Add choices for dropdown filters
        context['hazard_types'] = Hazard.HAZARD_TYPE_CHOICES
        context['risk_levels'] = Hazard.SEVERITY_CHOICES
        context['status_choices'] = HAZARD_STATUS_FILTER_CHOICES

        # Retain filter values in the form after submission
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_hazard_type'] = self.request.GET.get('hazard_type', '')
        context['selected_risk_level'] = self.request.GET.get('risk_level', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_assigned_by'] = self.request.GET.get('assigned_by', '')
        context['selected_assigned_to'] = selected_assigned_to

        return context
class HazardCreateView(LoginRequiredMixin, CreateView):
    model = Hazard
    form_class = HazardForm
    template_name = 'hazards/hazard_create.html'
    success_url = reverse_lazy('hazards:hazard_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        context['user_assigned_plants'] = user.assigned_plants.filter(is_active=True)

        context['departments'] = Department.objects.filter(is_active=True).order_by('name')
        return context

    def post(self, request, *args, **kwargs):
        """Override post to handle hazard submissions - bypasses form validation"""
        self.object = None
        # Always use handle_multiple_hazards since template uses custom field names
        return self.handle_multiple_hazards(request)
    
    def handle_multiple_hazards(self, request):
        """Handle single or multiple hazard submissions"""
        user = request.user
        hazard_count = int(request.POST.get('hazard_count', 1))
        
        created_hazards = []
        photos_uploaded_total = 0
        
        print(f"\n{'='*80}")
        print(f"🔄 Processing {hazard_count} hazard(s)")
        print(f"{'='*80}\n")
        print("FILES:", request.FILES)
        
        for hazard_index in range(hazard_count):
            print(f"\n--- Processing Hazard #{hazard_index + 1} ---")
            
            # Create new hazard instance
            hazard = Hazard()
            prefix = f'hazard_{hazard_index}_'
            
            # Reporter fields
            hazard.reported_by = user
            hazard.reporter_name = user.get_full_name()
            hazard.reporter_email = user.email
            hazard.reporter_phone = getattr(user, 'phone', '')
            # hazard.report_timestamp = timezone.now()
            hazard.report_source = 'web_portal'
            
            # Get hazard-specific fields
            hazard_type = request.POST.get(f'{prefix}hazard_type')
            hazard_category = request.POST.get(f'{prefix}hazard_category')
            severity = request.POST.get(f'{prefix}severity')
            hazard_description = request.POST.get(f'{prefix}hazard_description')
            immediate_action = request.POST.get(f'{prefix}immediate_action', '')
            
            print(f"  Type: {hazard_type}, Category: {hazard_category}, Severity: {severity}")
            
            # Validate required fields
            if not hazard_type or not hazard_category or not severity or not hazard_description:
                messages.error(request, f'Missing required fields for Hazard #{hazard_index + 1}')
                return redirect('hazards:hazard_create')
            
            hazard.hazard_type = hazard_type
            hazard.hazard_category = hazard_category
            hazard.severity = severity
            hazard.hazard_description = hazard_description
            hazard.immediate_action = immediate_action
            
            # Get location fields - try all possible variations
            plant_id = (
                request.POST.get('plant') or 
                request.POST.get('id_plant') or
                request.POST.get(f'{prefix}plant') or
                request.POST.get(f'{prefix}id_plant')
            )
            
            zone_id = (
                request.POST.get('zone') or 
                request.POST.get('id_zone') or
                request.POST.get(f'{prefix}zone') or
                request.POST.get(f'{prefix}id_zone')
            )
            
            location_id = (
                request.POST.get('location') or 
                request.POST.get('id_location') or
                request.POST.get(f'{prefix}location') or
                request.POST.get(f'{prefix}id_location')
            )
            
            sublocation_id = (
                request.POST.get('sublocation') or 
                request.POST.get('id_sublocation') or
                request.POST.get(f'{prefix}sublocation') or
                request.POST.get(f'{prefix}id_sublocation')
            )
            
            # Fallback to user's assigned locations
            if not plant_id and user.plant:
                plant_id = user.plant.id
            if not zone_id and hasattr(user, 'zone') and user.zone:
                zone_id = user.zone.id
            if not location_id and hasattr(user, 'location') and user.location:
                location_id = user.location.id
            
            print(f"  Plant: {plant_id}, Zone: {zone_id}, Location: {location_id}")
            
            # Validate required location fields
            if not plant_id:
                messages.error(request, f'Plant is required for Hazard #{hazard_index + 1}')
                return redirect('hazards:hazard_create')
            
            if not location_id:
                messages.error(request, f'Location is required for Hazard #{hazard_index + 1}')
                return redirect('hazards:hazard_create')
            
            hazard.plant_id = plant_id
            hazard.zone_id = zone_id if zone_id else None
            hazard.location_id = location_id
            hazard.sublocation_id = sublocation_id if sublocation_id else None
            
            # Incident datetime
            incident_datetime_str = request.POST.get('incident_datetime')
            if incident_datetime_str:
                try:
                    hazard.incident_datetime = datetime.datetime.fromisoformat(incident_datetime_str)
                    if timezone.is_naive(hazard.incident_datetime):
                        hazard.incident_datetime = timezone.make_aware(hazard.incident_datetime)
                except:
                    hazard.incident_datetime = timezone.now()
            else:
                hazard.incident_datetime = timezone.now()
            
            # On behalf logic
            behalf_checkbox = request.POST.get(f'{prefix}behalf_checkbox')
            if behalf_checkbox:
                hazard.behalf_person_name = request.POST.get(f'{prefix}behalf_person_name', '')
                behalf_dept_id = request.POST.get(f'{prefix}behalf_person_dept')
                if behalf_dept_id:
                    hazard.behalf_person_dept_id = behalf_dept_id
            
            # Title
            type_display = dict(Hazard.HAZARD_TYPE_CHOICES).get(hazard_type, hazard_type)
            category_display = dict(Hazard.HAZARD_CATEGORIES).get(hazard_category, hazard_category)
            hazard.hazard_title = f"{type_display} - {category_display}"
            
            # Status
            hazard.status = 'REPORTED'
            hazard.approval_status = 'PENDING'
            
            # Deadline
            severity_days = {'low': 30, 'medium': 15, 'high': 7, 'critical': 1}
            base_date = timezone.now().date()
            hazard.action_deadline = base_date + timezone.timedelta(
                days=severity_days.get(severity, 15)
            )
            
            # Save hazard
            try:
                hazard.save()
                print(f"  ✅ Saved with ID: {hazard.id}")
            except Exception as e:
                print(f"  ❌ Save error: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f'Error saving Hazard #{hazard_index + 1}: {str(e)}')
                continue
            
            # Generate report number
            # today = timezone.now().date()
            # plant_code = hazard.plant.code if hazard.plant else 'UNKN'
            # count = Hazard.objects.filter(created_at__date=today).count()
            # hazard.report_number = f"HAZ-{plant_code}-{today:%Y%m%d}-{count:03d}"
            # hazard.save(update_fields=['report_number'])
            # print(f"  📋 Report: {hazard.report_number}")
            
            # Handle photos
            photos_uploaded = 0
            photo_count = int(request.POST.get(f'{prefix}photo_count', 1))
            
            for i in range(photo_count):
                photo_key = f'{prefix}photo_{i}'
                photo = request.FILES.get(photo_key)

                if photo:
                    try:
                        compressed_photo = compress_image(photo)
                        HazardPhoto.objects.create(
                            hazard=hazard,
                            photo=compressed_photo,
                            photo_type='evidence',
                            uploaded_by=user
                        )
                        photos_uploaded += 1
                    except Exception as e:
                        print(f"Photo error: {e}")

            for video in request.FILES.getlist(f'{prefix}videos'):
                try:
                    compressed_video = compress_video(video)
                    HazardVideo.objects.create(
                        hazard=hazard,
                        video=compressed_video,
                        video_type='evidence',
                        uploaded_by=user
                    )
                except Exception as e:
                    print(f"Video error: {e}")
            
            photos_uploaded_total += photos_uploaded
            created_hazards.append(hazard)
            
            # Send notifications
            try:
                from apps.notifications.services import NotificationService
                NotificationService.notify(
                    content_object=hazard,
                    notification_type='HAZARD_REPORTED',
                    module='HAZARD'
                )
            except Exception as e:
                print(f"  Notification error: {e}")
        
        print(f"\n✅ Total created: {len(created_hazards)}")
        
        # Success messages
        if not created_hazards:
            messages.error(request, 'No hazards were created. Please try again.')
            return redirect('hazards:hazard_create')
        
        if len(created_hazards) == 1:
            hazard = created_hazards[0]
            messages.success(
                request,
                mark_safe(
                    f'<strong>✅ Hazard Report Submitted!</strong><br>'
                    f'Report No: {hazard.report_number}<br>'
                    f'Severity: {hazard.get_severity_display()}<br>'
                    f'Photos: {photos_uploaded_total}'
                )
            )
        else:
            report_numbers = ', '.join([h.report_number for h in created_hazards])
            messages.success(
                request,
                mark_safe(
                    f'<strong>✅ {len(created_hazards)} Hazards Submitted!</strong><br>'
                    f'Reports: {report_numbers}<br>'
                    f'Photos: {photos_uploaded_total}'
                )
            )
        
        return redirect(self.success_url)

    
    
class HazardDetailView(LoginRequiredMixin, DetailView):
    """
    Display details of a specific hazard, optimized for performance.
    """
    model = Hazard
    template_name = 'hazards/hazard_detail.html'
    context_object_name = 'hazard'

    def get_queryset(self):
        """
        Optimize the query by pre-fetching related objects to avoid
        multiple database hits in the template.
        """
       
        return Hazard.objects.select_related(
            'plant', 'zone', 'location', 'sublocation',
            'reported_by', 'assigned_to', 'approved_by',
            'behalf_person', 
            'behalf_person_dept'
        ).prefetch_related(
            'photos', 
            'videos',
            'action_items'  # ✅ FIXED: Just prefetch action_items (no responsible_person)
        )

    def get_context_data(self, **kwargs):
        """
        Add the prefetched photos and action items to the context so the
        template can access them directly.
        """
        
        context = super().get_context_data(**kwargs)
        
        hazard = self.get_object()
        context['action_items'] = hazard.action_items.all()
        context['photos'] = hazard.photos.all()
        context['videos'] = hazard.videos.all()
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        
        return context

class HazardUpdateView(LoginRequiredMixin, UpdateView):
    model = Hazard
    form_class = HazardForm
    template_name = 'hazards/hazard_update.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_success_url(self):
        return reverse_lazy('hazards:hazard_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        # Get user's assigned locations for autofill logic
        context['user_assigned_plants'] = user.assigned_plants.filter(is_active=True)
        
        # Get the current hazard
        hazard = self.object
        
        # Get departments for behalf dropdown
        context['departments'] = Department.objects.filter(is_active=True).order_by('name')
        context['photos'] = hazard.photos.all()
        context['videos'] = hazard.videos.all()
        return context

    def form_valid(self, form):
        hazard = form.save(commit=False)
        user = self.request.user

        # Update reporter fields
        hazard.reporter_name = user.get_full_name()
        hazard.reporter_email = user.email
        hazard.reporter_phone = getattr(user, 'phone', '')
        
        # Update title based on type and category
        type_display = dict(Hazard.HAZARD_TYPE_CHOICES).get(hazard.hazard_type, hazard.hazard_type)
        category_display = dict(Hazard.HAZARD_CATEGORIES).get(hazard.hazard_category, hazard.hazard_category)
        hazard.hazard_title = f"{type_display} - {category_display}"
        
        # Update deadline based on severity
        severity_days = {'low': 30, 'medium': 15, 'high': 7, 'critical': 1}
        base_date = hazard.incident_datetime.date() if hazard.incident_datetime else timezone.now().date()
        hazard.action_deadline = base_date + timezone.timedelta(
            days=severity_days.get(hazard.severity, 15)
        )
        
        # Handle behalf logic
        behalf_checkbox = self.request.POST.get('behalf_checkbox')
        if behalf_checkbox:
            hazard.behalf_person_name = self.request.POST.get('behalf_person_name', '')
            behalf_dept_id = self.request.POST.get('behalf_person_dept')
            if behalf_dept_id:
                hazard.behalf_person_dept_id = behalf_dept_id
        else:
            hazard.behalf_person_name = None
            hazard.behalf_person_dept = None
        
        # Save the hazard
        hazard.save()
        print(f"✅ Hazard updated: {hazard.report_number}")
        
        # Handle photo deletion
        for key in self.request.POST:
            if key.startswith('keep_photo_') and self.request.POST[key] == '0':
                photo_id = key.split('_')[-1]
                HazardPhoto.objects.filter(id=photo_id, hazard=hazard).delete()
                print(f"🗑️ Deleted photo: {photo_id}")

        # Handle new photo uploads
        photo_index = 0
        while True:
            photo_key = f'photo_{photo_index}'
            if photo_key in self.request.FILES:
                try:
                    photo = self.request.FILES[photo_key]
                    compressed_photo = compress_image(photo)
                    HazardPhoto.objects.create(
                        hazard=hazard,
                        photo=compressed_photo,
                        photo_type='evidence',
                        uploaded_by=user
                    )
                    print(f"📸 Added new photo: {photo_key}")
                except Exception as e:
                    print(f"❌ Error uploading photo {photo_key}: {e}")
                photo_index += 1
            else:
                break

        for video in self.request.FILES.getlist('videos'):
            try:
                compressed_video = compress_video(video)
                HazardVideo.objects.create(
                    hazard=hazard,
                    video=compressed_video,
                    video_type='evidence',
                    uploaded_by=user
                )
                print("Added new video")
            except Exception as e:
                print(f"Error uploading video: {e}")
        
        # Success message
        messages.success(
            self.request,
            mark_safe(
                f'<strong>✅ Hazard Report Updated!</strong><br>'
                f'Report No: {hazard.report_number}<br>'
                f'Severity: {hazard.get_severity_display()}'
            )
        )
        
        print(f"\n✅ Update complete for {hazard.report_number}")
        print("="*80 + "\n")
        
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        """Handle form validation errors"""
        for field, errors in form.errors.items():
            print(f"  {field}: {errors}")
        return super().form_invalid(form)

class HazardDeleteView(LoginRequiredMixin, View):

    def get(self, request, pk):
        hazard = get_object_or_404(Hazard, pk=pk)

        return render(
            request,
            'hazards/hazard_confirm_delete.html',
            {'hazard': hazard}
        )

    def post(self, request, pk):
        hazard = get_object_or_404(Hazard, pk=pk)

        hazard.delete()

        messages.success(request, "Hazard deleted successfully.")

        return redirect('hazards:hazard_list')
                 
class HazardActionItemCreateView(LoginRequiredMixin, CreateView):
    """
    Create an action item for a specific hazard.
    Handles form submission for creating new HazardActionItem(s),
    including file attachments.
    """
    model = HazardActionItem
    template_name = 'hazards/action_item_create.html'
    fields = []

    def dispatch(self, request, *args, **kwargs):
        """Ensure the hazard exists before proceeding."""
        self.hazard = get_object_or_404(Hazard, pk=self.kwargs['hazard_pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hazard'] = self.hazard

        # Calculate auto target date (no changes here)
        if hasattr(self.hazard, 'get_severity_deadline_days'):
            severity_days = self.hazard.get_severity_deadline_days()
        else:
            severity_map = {'low': 30, 'medium': 15, 'high': 7, 'critical': 1}
            severity_days = severity_map.get(self.hazard.severity, 15)

        auto_target_date = timezone.now().date() + timezone.timedelta(days=severity_days)
        context['auto_target_date'] = auto_target_date.strftime('%Y-%m-%d')
        context['severity_days'] = severity_days
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')

        user = self.request.user

        # ===================================================================
        # START OF THE FIX
        # ===================================================================
        
        # Initialize defaults
        context['plant_users'] = []
        context['plant_name'] = 'the relevant location'

        # Fetch users only if the hazard has a plant assigned
        if self.hazard.plant:
            from django.db.models import Q

            # Step 1: Start by filtering users based on the hazard's plant
            plant_filter = Q(plant=self.hazard.plant) | Q(assigned_plants=self.hazard.plant)
            responsible_users_query = User.objects.filter(plant_filter)

            # Step 2: If the hazard also has a zone, add a zone filter to the query
            if self.hazard.zone:
                zone_filter = Q(zone=self.hazard.zone) | Q(assigned_zones=self.hazard.zone)
                responsible_users_query = responsible_users_query.filter(zone_filter)
                # Update the display name to include the zone
                context['plant_name'] = f"{self.hazard.plant.name} - {self.hazard.zone.name}"
            else:
                # If no zone, just use the plant name
                context['plant_name'] = self.hazard.plant.name

            # Step 3: Apply final filters (active status, exclude self, etc.)
            final_users = responsible_users_query.filter(
                is_active=True,
                is_active_employee=True
            ).exclude(
                id=user.id
            ).distinct().select_related(
                'department', 'role'
            ).order_by('first_name', 'last_name')
            
            context['plant_users'] = final_users
            
        # ===================================================================
        # END OF THE FIX
        # ===================================================================

        return context

    def post(self, request, *args, **kwargs):
        """Handle the POST request to create new action item(s)."""

        # print("\n" + "=" * 80)
        # print("🎯 ACTION ITEM CREATION")
        # print("=" * 80)

        assignment_type = request.POST.get('assignment_type')
        # print(f"Assignment Type: {assignment_type}")

        try:
            action_description = request.POST.get('action_description', '').strip()
            target_date_str = request.POST.get('target_date')
            attachment = request.FILES.get('attachment')
            
            if assignment_type == 'self' and not action_description:
                messages.error(request, 'The "Action Taken" description is required for self-assignment.')
                return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

            if not target_date_str:
                messages.error(request, 'Target date is required.')
                return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

            target_date = datetime.datetime.strptime(target_date_str, '%Y-%m-%d').date()

            # ============================================================
            # ✅ SELF ASSIGNMENT (Complete immediately + Close hazard)
            # ============================================================
            if assignment_type == 'self':

                if not attachment:
                    messages.error(request, 'An attachment is required to self-assign and close an action.')
                    return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

                # --- REFACTORED LOGIC ---
                # Step 1: Create the instance with initial data, but don't handle M2M yet.
                # Status starts as 'PENDING' and will be updated by the logic in the save method.
                action_item = HazardActionItem(
                    hazard=self.hazard,
                    action_description=action_description,
                    created_by=request.user,
                    is_self_assigned=True,
                    target_date=target_date,
                    responsible_emails=request.user.email,
                    status='PENDING', # Will be updated after adding the user
                    completion_remarks='Self-assigned and completed by reporter.',
                    attachment=attachment
                )
                action_item.save()  # First save to get an ID.

                # Step 2: Now that it has an ID, add the user to the M2M relationship.
                action_item.completed_by_users.add(request.user)

                # Step 3: Save again. Now the model's save() method will correctly detect
                # that it is fully completed and update the status to 'COMPLETED'.
                action_item.save()

                # print(f"✅ Self-assigned to: {request.user.email}")
                # print(f"💾 Action item ID: {action_item.id}")
                # --- END OF REFACTORED LOGIC ---

                # Close hazard
                self.hazard.status = 'CLOSED'
                self.hazard.save(update_fields=['status'])
                # print("🔒 Hazard status updated to: CLOSED")

                # ... (success message remains the same) ...

            # ============================================================
            # ✅ FORWARD ASSIGNMENT (Create one per selected user)
            # ============================================================
            else:
                # This part now works correctly because of the fix in the model's save method.
                # No changes are needed here.
                responsible_emails = request.POST.getlist('responsible_emails')
                print(f"📧 Responsible emails from POST: {responsible_emails}")

                if not responsible_emails:
                    messages.error(request, 'Please select at least one user to assign this action to.')
                    return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

                responsible_emails_str = ",".join(responsible_emails)

                action_item = HazardActionItem.objects.create(
                    hazard=self.hazard,
                    action_description=action_description,
                    created_by=request.user,
                    is_self_assigned=False,
                    target_date=target_date,
                    responsible_emails=responsible_emails_str,
                    status='PENDING',
                    attachment=attachment
                )

                print(f"💾 1 action item created for {len(responsible_emails)} user(s).")
                # Update hazard status
                self.hazard.status = 'ACTION_ASSIGNED'
                self.hazard.save(update_fields=['status'])
                print("📋 Hazard status updated to: ACTION_ASSIGNED")

                # Send notifications to all selected users
                try:
                    responsible_users = User.objects.filter(
                        email__in=responsible_emails,
                        is_active=True
                    )

                    NotificationService.notify(
                        content_object=action_item, # Use the single created item for notification
                        notification_type='HAZARD_ACTION_ASSIGNED',
                        module='HAZARD_ACTION',
                        extra_recipients=list(responsible_users)
                    )

                except Exception as e:
                    print(f"⚠️ Notification error: {e}")

                # Update the success message for a single action item
                message = mark_safe(
                    f'✅ <strong>Action Item Created!</strong><br>'
                    f'Assigned to <strong>{len(responsible_emails)}</strong> user(s)<br>'
                    f'Status: <strong>PENDING</strong><br>'
                    f'Hazard status: <strong>ACTION_ASSIGNED</strong>'
                )
                messages.success(request, message)

            print("=" * 80 + "\n")
            return redirect('hazards:hazard_detail', pk=self.hazard.pk)

        except Exception as e:
            print(f"❌ Error creating action item: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error creating action item: {str(e)}')
            return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

    def get_success_url(self):
        """Redirect to hazard detail page on success."""
        return reverse_lazy('hazards:hazard_detail', kwargs={'pk': self.hazard.pk})
    

class HazardActionItemUpdateView(LoginRequiredMixin, UpdateView):
    model = HazardActionItem
    template_name = "hazards/action_item_update.html"
    fields = []

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.hazard = self.object.hazard
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["hazard"] = self.hazard
        user = self.request.user

        # ===================================================================
        # START OF THE FIX
        # ===================================================================
        
        # Initialize defaults
        context['plant_users'] = []
        context['plant_name'] = 'the relevant location'

        # Fetch users based on hazard's plant and zone (copied from CreateView)
        if self.hazard.plant:
            from django.db.models import Q

            # Step 1: Start by filtering users based on the hazard's plant
            plant_filter = Q(plant=self.hazard.plant) | Q(assigned_plants=self.hazard.plant)
            responsible_users_query = User.objects.filter(plant_filter)

            # Step 2: If the hazard also has a zone, add a zone filter to the query
            if self.hazard.zone:
                zone_filter = Q(zone=self.hazard.zone) | Q(assigned_zones=self.hazard.zone)
                responsible_users_query = responsible_users_query.filter(zone_filter)
                # Update the display name to include the zone
                context['plant_name'] = f"{self.hazard.plant.name} - {self.hazard.zone.name}"
            else:
                # If no zone, just use the plant name
                context['plant_name'] = self.hazard.plant.name

            # Step 3: Apply final filters (active status, exclude self, etc.)
            final_users = responsible_users_query.filter(
                is_active=True,
                is_active_employee=True
            ).exclude(
                id=user.id
            ).distinct().select_related(
                'department', 'role'
            ).order_by('first_name', 'last_name')
            
            context['plant_users'] = final_users
        
        # ===================================================================
        # END OF THE FIX
        # ===================================================================
        
        # Keep the rest of the original context data logic
        selected_emails = self.object.get_emails_list()
        selected_users = User.objects.filter(
            email__in=selected_emails
        ).select_related('department', 'role')

        context['selected_users'] = selected_users
        context['selected_emails'] = selected_emails

        context["severity_days"] = (
            (self.hazard.action_deadline - self.hazard.incident_datetime.date()).days
            if self.hazard.action_deadline
            else 0
        )

        context["auto_target_date"] = (
            self.object.target_date.strftime("%Y-%m-%d")
            if self.object.target_date
            else ""
        )

        context["cancel_url"] = (
            self.request.GET.get("next")
            or self.request.META.get("HTTP_REFERER")
            or "/"
        )
        context["is_self_assignment"] = self.object.is_self_assigned

        return context
    
    def get_success_url(self):
        return reverse_lazy(
            "hazards:hazard_detail",
            kwargs={"pk": self.object.hazard.pk},
        )

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        try:
            assignment_type = request.POST.get("assignment_type")
            action_description = request.POST.get("action_description", "").strip()

            if assignment_type == 'self' and not action_description:
                messages.error(request, 'The "Action Taken" description is required for self-assignment.')
                return redirect('hazards:action_item_update', pk=self.object.pk)

            self.object.action_description = action_description

            target_date = request.POST.get("target_date")
            if target_date:
                self.object.target_date = datetime.datetime.strptime(
                    target_date, "%Y-%m-%d"
                ).date()

            if "attachment" in request.FILES:
                self.object.attachment = request.FILES["attachment"]

            if assignment_type == "self":
                # Attachment is required for self-completion
                if "attachment" not in request.FILES and not self.object.attachment:
                    messages.error(
                        request,
                        "An attachment is required for self-assignment and completion.",
                    )
                    return redirect(
                        "hazards:action_item_update",
                        pk=self.object.pk,
                    )

                # Update assignment details
                self.object.responsible_emails = request.user.email
                self.object.is_self_assigned = True
                self.object.save()  # Save before changing M2M

                # CRITICAL FIX: Reset completion and then mark as complete by current user
                self.object.completed_by_users.clear()
                self.object.completed_by_users.add(request.user)

            elif assignment_type == "forward":
                selected_emails = request.POST.getlist("responsible_emails")

                if not selected_emails:
                    messages.error(
                        request,
                        "Please select at least one user.",
                    )
                    return redirect(
                        "hazards:action_item_update",
                        pk=self.object.pk,
                    )

                # Update assignment details
                self.object.responsible_emails = ",".join(selected_emails)
                self.object.is_self_assigned = False
                self.object.save() # Save before changing M2M

                # CRITICAL FIX: Clear all previous completions as assignees have changed
                self.object.completed_by_users.clear()

            # Final save will trigger the model's logic to set the correct status
            self.object.save()

            # Update parent hazard status
            if hasattr(self.object.hazard, "update_status_from_action_items"):
                self.object.hazard.update_status_from_action_items()

            messages.success(
                request,
                mark_safe(
                    f"✅ <strong>Action item updated successfully!</strong><br>"
                    f"Status: <strong>{self.object.get_status_display()}</strong>"
                ),
            )

            return redirect(self.get_success_url())

        except Exception as e:
            print("Update Error:", e)
            messages.error(
                request,
                f"Error updating action item: {str(e)}",
            )
            return redirect(
                "hazards:action_item_update",
                pk=self.object.pk,
            )


class HazardDashboardViews(LoginRequiredMixin, TemplateView):
    """
    Advanced Hazard Management Dashboard with working filters.
    """
    template_name = 'hazards/hazards_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = datetime.date.today()

        # 1. Get all filter parameters from the URL
        selected_plant = self.request.GET.get('plant', '')
        selected_zone = self.request.GET.get('zone', '')
        selected_location = self.request.GET.get('location', '')
        selected_sublocation = self.request.GET.get('sublocation', '')
        selected_month = self.request.GET.get('month', '')
        selected_severity = self.request.GET.get('severity', '')
        selected_status = self.request.GET.get('status', '')
        selected_overdue = self.request.GET.get('overdue', '')
        selected_closed = self.request.GET.get('closed', '')
        selected_hazard_type = self.request.GET.get('hazard_type', '')
        selected_category = self.request.GET.get('category', '')    # <-- NEW
        selected_department = self.request.GET.get('department', '') # <-- NEW

        # 2. Build the base queryset based on user role
        user_plants = Plant.objects.none()
        if user.is_superuser or getattr(user, 'role', None) and user.role.name == 'ADMIN':
            base_hazards = Hazard.objects.all()
            user_plants = Plant.objects.filter(is_active=True).order_by('name')
        elif user.get_all_plants():
            user_plants = user.get_all_plants()
            base_hazards = Hazard.objects.filter(plant__in=user_plants).distinct()
        else:
            base_hazards = Hazard.objects.filter(reported_by=user)
            user_plants = Plant.objects.none()

        # 3. Calculate top-level stats BEFORE applying any filters.
        # context['total_hazards'] = base_hazards.count()
        # context['closed_hazards_count'] = base_hazards.filter(status='CLOSED').count()
        # context['overdue_hazards_count'] = base_hazards.filter(action_deadline__lt=today).exclude(status='CLOSED').count()
        # this_month_total = base_hazards.filter(incident_datetime__year=today.year, incident_datetime__month=today.month).count()


        # 4. Apply filters to a new queryset for charts and lists.
        filtered_hazards = base_hazards
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                filtered_hazards = filtered_hazards.filter(
                    incident_datetime__year=year,
                    incident_datetime__month=month
                )
            except (ValueError, TypeError):
                pass
        if selected_plant:
            filtered_hazards = filtered_hazards.filter(plant_id=selected_plant)
        if selected_zone:
            filtered_hazards = filtered_hazards.filter(zone_id=selected_zone)
        if selected_location:
            filtered_hazards = filtered_hazards.filter(location_id=selected_location)
        if selected_sublocation:
            filtered_hazards = filtered_hazards.filter(sublocation_id=selected_sublocation)
        if selected_severity:
            filtered_hazards = filtered_hazards.filter(severity=selected_severity)
        if selected_hazard_type:
            filtered_hazards = filtered_hazards.filter(hazard_type=selected_hazard_type)
        if selected_category: # <-- NEW
            filtered_hazards = filtered_hazards.filter(hazard_category=selected_category)
        if selected_department: # <-- NEW
            filtered_hazards = filtered_hazards.filter(
                Q(assigned_to__department_id=selected_department) |
                Q(assigned_to__department__isnull=True, reported_by__department_id=selected_department)
            )
            
        filtered_hazards = filter_hazards_by_status(filtered_hazards, selected_status)
                
        if selected_overdue == 'true':
            filtered_hazards = filter_hazards_by_status(filtered_hazards, 'OVERDUE')

        if selected_closed == 'true':
            filtered_hazards = filter_hazards_by_status(filtered_hazards, 'CLOSED')
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                filtered_hazards = filtered_hazards.filter(incident_datetime__year=year, incident_datetime__month=month)
            except (ValueError, TypeError):
                pass
        
        # context['this_month_hazards'] = filtered_hazards.count() if selected_month else this_month_total
        # context['current_month_value'] = today.strftime('%Y-%m')
        context['current_month_value'] = selected_month if selected_month else today.strftime('%Y-%m')

        # Dashboard cards should respect active filters

        if today.month >= 4:  # Apr-Dec
            fy_start = datetime.date(today.year, 4, 1)
            fy_end = datetime.date(today.year + 1, 3, 31)
        else:  # Jan-Mar
            fy_start = datetime.date(today.year - 1, 4, 1)
            fy_end = datetime.date(today.year, 3, 31)

        context['total_hazards'] = base_hazards.filter(
            incident_datetime__date__gte=fy_start,
            incident_datetime__date__lte=fy_end
        ).count()

        context['closed_hazards_count'] = filter_hazards_by_status(
            filtered_hazards,
            'CLOSED'
        ).count()

        context['overdue_hazards_count'] = filter_hazards_by_status(
            filtered_hazards,
            'OVERDUE'
        ).count()

        context['this_month_hazards'] = filtered_hazards.count()

        # 5. Prepare filter dropdown options
        context['plants'] = user_plants
        
        zone_qs = Zone.objects.filter(is_active=True)
        if not user.is_superuser and not (getattr(user, 'role', None) and user.role.name == 'ADMIN'):
            zone_qs = zone_qs.filter(plant__in=user_plants)
        if selected_plant:
            zone_qs = zone_qs.filter(plant_id=selected_plant)
        context['zones'] = zone_qs.order_by('name')

        location_qs = Location.objects.filter(is_active=True)
        if not user.is_superuser and not (getattr(user, 'role', None) and user.role.name == 'ADMIN'):
            location_qs = location_qs.filter(zone__plant__in=user_plants)
        if selected_zone:
            location_qs = location_qs.filter(zone_id=selected_zone)
        elif selected_plant:
            location_qs = location_qs.filter(zone__plant_id=selected_plant)

        context['locations'] = location_qs.order_by('name')

        sublocation_qs = SubLocation.objects.filter(is_active=True)
        if not user.is_superuser and not (getattr(user, 'role', None) and user.role.name == 'ADMIN'):
            sublocation_qs = sublocation_qs.filter(location__zone__plant__in=user_plants)
        if selected_location:
            sublocation_qs = sublocation_qs.filter(location_id=selected_location)
        elif selected_zone:
            sublocation_qs = sublocation_qs.filter(location__zone_id=selected_zone)
        elif selected_plant:
            sublocation_qs = sublocation_qs.filter(location__zone__plant_id=selected_plant)

        context['sublocations'] = sublocation_qs.order_by('name')
        
        context['month_options'] = [{
            'value': (today - datetime.timedelta(days=i*30)).strftime('%Y-%m'),
            'label': (today - datetime.timedelta(days=i*30)).strftime('%B %Y')
        } for i in range(12)]

        context['all_departments'] = Department.objects.filter(is_active=True).order_by('name') # <-- NEW
        context['hazard_types'] = Hazard.HAZARD_TYPE_CHOICES
        context['all_categories'] = Hazard.HAZARD_CATEGORIES # <-- NEW
        context['status_choices'] = HAZARD_STATUS_FILTER_CHOICES

        # Pass selected filter values and names back to the template
        context.update({
            'selected_plant': selected_plant, 'selected_zone': selected_zone,
            'selected_location': selected_location, 'selected_sublocation': selected_sublocation,
            'selected_month': selected_month, 'selected_severity': selected_severity,
            'selected_status': selected_status,
            'selected_hazard_type': selected_hazard_type,
            'selected_category': selected_category, # <-- NEW
            'selected_department': selected_department, # <-- NEW
            'selected_overdue': selected_overdue,
            'selected_closed': selected_closed,
        })
        try:
            if selected_plant: context['selected_plant_name'] = Plant.objects.get(id=selected_plant).name
            if selected_zone: context['selected_zone_name'] = Zone.objects.get(id=selected_zone).name
            if selected_location: context['selected_location_name'] = Location.objects.get(id=selected_location).name
            if selected_sublocation: context['selected_sublocation_name'] = SubLocation.objects.get(id=selected_sublocation).name
            if selected_department: context['selected_department_name'] = Department.objects.get(id=selected_department).name # <-- NEW
            if selected_hazard_type:context['selected_hazard_type_name'] = dict(Hazard.HAZARD_TYPE_CHOICES).get(selected_hazard_type)
            if selected_category: context['selected_category_name'] = dict(Hazard.HAZARD_CATEGORIES).get(selected_category) # <-- NEW
            if selected_status: context['selected_status_label'] = dict(HAZARD_STATUS_FILTER_CHOICES).get(selected_status, selected_status.replace('_', ' ').title())
            if selected_month:
                year, month = map(int, selected_month.split('-'))
                context['selected_month_label'] = datetime.date(year, month, 1).strftime('%B %Y')
        except:
             pass
        context['has_active_filters'] = any(context.get(key) for key in ['selected_plant', 'selected_zone', 'selected_location', 'selected_sublocation', 'selected_month', 'selected_severity', 'selected_status', 'selected_category', 'selected_department', 'selected_overdue', 'selected_closed'])
        # 6. Prepare data for lists and charts using the FILTERED queryset
        # context['recent_hazards'] = filtered_hazards.select_related('plant', 'location').order_by('-incident_datetime')[:10]

        from django.core.paginator import Paginator

        hazards_qs = filtered_hazards.select_related(
            'plant',
            'location',
            'reported_by',
            'reported_by__department',
            'assigned_to',
            'assigned_to__department',
        ).prefetch_related(
            'action_items',
            'action_items__completed_by_users',
        ).order_by('-incident_datetime')

        paginator = Paginator(hazards_qs, 10)  # 10 per page
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['page_obj'] = page_obj
        context['is_paginated'] = page_obj.has_other_pages()
        context['recent_hazards'] = page_obj.object_list

        from urllib.parse import urlencode

        querydict = self.request.GET.copy()
        querydict.pop('page', None)

        context['querystring'] = urlencode(querydict)
        context['current_filters'] = querydict.urlencode()

        # --- PIE CHART DATA ---
        top_categories_query = filtered_hazards.values('hazard_category').annotate(count=Count('hazard_category')).order_by('-count')[:3]
        category_display_map = dict(Hazard.HAZARD_CATEGORIES)
        
        # Data for Pie Chart JavaScript
        top_category_labels, top_category_data, top_category_values = [], [], []
        for item in top_categories_query:
            top_category_labels.append(category_display_map.get(item['hazard_category'], 'Unknown'))
            top_category_data.append(item['count'])
            top_category_values.append(item['hazard_category'])
            
        context['top_category_labels'] = json.dumps(top_category_labels)
        context['top_category_data'] = json.dumps(top_category_data)
        context['top_category_values'] = json.dumps(top_category_values)
        
        # ✅ *** THE CRITICAL FIX IS HERE ***
        # We must pass the actual data (the queryset) to the template, NOT a boolean.
        context['top_hazard_categories'] = top_categories_query  
        # context['top_hazard_categories'] = top_categories_query.exists() # Check if data exists for chart

        # ... (Monthly Trend, Severity, and Status charts ka code waise hi rahega) ...
        # Monthly Trend ...
        six_months_ago = today - datetime.timedelta(days=180)
        monthly_hazards = filtered_hazards.filter(incident_datetime__gte=six_months_ago).annotate(month=TruncMonth('incident_datetime')).values('month').annotate(count=Count('id')).order_by('month')
        context['monthly_labels'] = json.dumps([item['month'].strftime('%b %Y') for item in monthly_hazards])
        context['monthly_data'] = json.dumps([item['count'] for item in monthly_hazards])

        # Severity Distribution ...
        severity_distribution = filtered_hazards.values('severity').annotate(count=Count('id'))
        severity_dict = {item['severity']: item['count'] for item in severity_distribution}
        severity_labels = [choice[1] for choice in Hazard.SEVERITY_CHOICES]
        severity_values = [choice[0] for choice in Hazard.SEVERITY_CHOICES]
        context['severity_labels'] = json.dumps(severity_labels)
        context['severity_data'] = json.dumps([severity_dict.get(val, 0) for val in severity_values])

        # Status Distribution ...
        status_counts = {}
        status_labels, status_keys, status_data = [], [], []
        status_choices_dict = dict(HAZARD_STATUS_FILTER_CHOICES)
        for hazard in filtered_hazards.prefetch_related('action_items'):
            status_counts[hazard.effective_status] = status_counts.get(hazard.effective_status, 0) + 1

        for status_key, count in sorted(status_counts.items(), key=lambda item: item[1], reverse=True):
            status_labels.append(status_choices_dict.get(status_key, status_key.replace('_', ' ').title()))
            status_keys.append(status_key)
            status_data.append(count)
        context['status_labels'] = json.dumps(status_labels)
        context['status_keys'] = json.dumps(status_keys)
        context['status_data'] = json.dumps(status_data)
        
        department_map = {}
        for hazard in hazards_qs:
            assigned_users = hazard.get_assigned_users()
            if assigned_users:
                hazard_department_keys = set()
                for assigned_user in assigned_users:
                    assigned_department = getattr(assigned_user, 'department', None)
                    department_key = assigned_department.id if assigned_department else 'unassigned'
                    if department_key in hazard_department_keys:
                        continue
                    hazard_department_keys.add(department_key)

                    if department_key not in department_map:
                        department_map[department_key] = {
                            'assigned_department_id': assigned_department.id if assigned_department else '',
                            'name': assigned_department.name if assigned_department else 'Unassigned',
                            'total': 0,
                            'closed_count': 0,
                            'overdue_count': 0,
                            'critical_count': 0,
                        }

                    row = department_map[department_key]
                    row['total'] += 1
                    if hazard.status == 'CLOSED':
                        row['closed_count'] += 1
                    elif hazard.action_deadline and hazard.action_deadline < today:
                        row['overdue_count'] += 1
                    if hazard.severity == 'critical':
                        row['critical_count'] += 1
            else:
                department_key = 'unassigned'
                if department_key not in department_map:
                    department_map[department_key] = {
                        'assigned_department_id': '',
                        'name': 'Unassigned',
                        'total': 0,
                        'closed_count': 0,
                        'overdue_count': 0,
                        'critical_count': 0,
                    }

                row = department_map[department_key]
                row['total'] += 1
                if hazard.status == 'CLOSED':
                    row['closed_count'] += 1
                elif hazard.action_deadline and hazard.action_deadline < today:
                    row['overdue_count'] += 1
                if hazard.severity == 'critical':
                    row['critical_count'] += 1

        department_distribution = sorted(
            department_map.values(),
            key=lambda item: (-item['total'], item['name'])
        )

        for item in department_distribution:
            item['open_count'] = item['total'] - item['closed_count']

        department_ids = [item['assigned_department_id'] for item in department_distribution]
        department_labels = [item['name'] for item in department_distribution]
        department_total_data = [item['total'] for item in department_distribution]
        department_open_data = [item['open_count'] for item in department_distribution]
        department_closed_data = [item['closed_count'] for item in department_distribution]
        department_overdue_data = [item['overdue_count'] for item in department_distribution]

        top_department = department_distribution[0] if department_distribution else None

        context['department_ids'] = json.dumps(department_ids)
        context['department_labels'] = json.dumps(department_labels)
        context['department_total_data'] = json.dumps(department_total_data)
        context['department_open_data'] = json.dumps(department_open_data)
        context['department_closed_data'] = json.dumps(department_closed_data)
        context['department_overdue_data'] = json.dumps(department_overdue_data)
        context['department_chart_data'] = bool(department_distribution)
        context['department_breakdown'] = department_distribution[:6]
        context['department_summary'] = {
            'departments_covered': len(department_distribution),
            'top_department_name': top_department['name'] if top_department else 'N/A',
            'top_department_total': top_department['total'] if top_department else 0,
            'top_department_open': top_department['open_count'] if top_department else 0,
        }

        # =====================================
        # Plant Wise Hazard Status Tracker
        # =====================================
        plant_summary_map = {}
        hazard_status_qs = filtered_hazards.select_related('plant').prefetch_related('action_items')

        for hazard in hazard_status_qs:
            plant_obj = hazard.plant
            plant_key = plant_obj.id if plant_obj else 'unassigned'
            plant_name = plant_obj.name if plant_obj else 'Unassigned'

            if plant_key not in plant_summary_map:
                plant_summary_map[plant_key] = {
                    'plant_id': plant_obj.id if plant_obj else '',
                    'plant_name': plant_name,
                    'closed_count': 0,
                    'pending_count': 0,
                    'in_progress_count': 0,
                    'cancelled_count': 0,
                    'overdue_count': 0,
                    'late_close_count': 0,
                    'total': 0,
                }

            row = plant_summary_map[plant_key]
            row['total'] += 1

            effective_status = hazard.effective_status
            if effective_status == 'CLOSED_LATE':
                row['late_close_count'] += 1
            elif effective_status == 'OVERDUE':
                row['overdue_count'] += 1
            elif effective_status == 'CLOSED':
                row['closed_count'] += 1
            elif effective_status == 'IN_PROGRESS':
                row['in_progress_count'] += 1
            elif effective_status == 'REJECTED':
                row['cancelled_count'] += 1
            else:
                row['pending_count'] += 1

        plant_summary = sorted(
            plant_summary_map.values(),
            key=lambda item: (-item['total'], item['plant_name'])
        )

        top_plant = plant_summary[0] if plant_summary else None
        plant_status_rows = []
        for item in plant_summary:
            total_count = item['total'] or 0
            closed_percent = round((item['closed_count'] / total_count) * 100) if total_count else 0
            plant_status_rows.append({
                'plant_id': item['plant_id'],
                'plant_name': item['plant_name'],
                'closed_count': item['closed_count'],
                'pending_count': item['pending_count'],
                'in_progress_count': item['in_progress_count'],
                'cancelled_count': item['cancelled_count'],
                'overdue_count': item['overdue_count'],
                'late_close_count': item['late_close_count'],
                'total': total_count,
                'closed_percent': closed_percent,
            })

        context['plant_chart_labels'] = json.dumps([item['plant_name'] for item in plant_summary])
        context['plant_chart_ids'] = json.dumps([item['plant_id'] for item in plant_summary])
        context['plant_closed_data'] = json.dumps([item['closed_count'] for item in plant_summary])
        context['plant_pending_data'] = json.dumps([item['pending_count'] for item in plant_summary])
        context['plant_in_progress_data'] = json.dumps([item['in_progress_count'] for item in plant_summary])
        context['plant_cancelled_data'] = json.dumps([item['cancelled_count'] for item in plant_summary])
        context['plant_overdue_data'] = json.dumps([item['overdue_count'] for item in plant_summary])
        context['plant_late_close_data'] = json.dumps([item['late_close_count'] for item in plant_summary])
        context['plant_chart_data'] = bool(plant_summary)
        context['plant_status_rows'] = plant_status_rows
        context['plant_summary'] = {
            'plants_covered': len(plant_summary),
            'top_plant_name': top_plant['plant_name'] if top_plant else 'N/A',
            'top_plant_total': top_plant['total'] if top_plant else 0,
            'top_plant_overdue': top_plant['overdue_count'] if top_plant else 0,
        }
        # Hazard Type Chart Data
        hazard_type_distribution = (
            filtered_hazards
            .values('hazard_type')
            .annotate(count=Count('id'))
        )

        hazard_type_dict = {
            item['hazard_type']: item['count']
            for item in hazard_type_distribution
        }

        hazard_type_labels = []
        hazard_type_counts = []
        hazard_type_values = []

        for value, label in Hazard.HAZARD_TYPE_CHOICES:
            hazard_type_labels.append(label)
            hazard_type_counts.append(hazard_type_dict.get(value, 0))
            hazard_type_values.append(value)

        context['hazard_type_labels'] = json.dumps(hazard_type_labels)
        context['hazard_type_counts'] = json.dumps(hazard_type_counts)
        context['hazard_type_values'] = json.dumps(hazard_type_values)


        # =====================================
        # Hazard Overdue By Department
        # =====================================

        overdue_department_map = {}

        overdue_hazards = filtered_hazards.filter(
            action_deadline__lt=today
        ).exclude(
            status='CLOSED'
        )

        for hazard in overdue_hazards:

            assigned_users = hazard.get_assigned_users()

            if assigned_users:

                for assigned_user in assigned_users:

                    dept = getattr(
                        assigned_user,
                        'department',
                        None
                    )

                    if dept:

                        overdue_department_map[dept.name] = (
                            overdue_department_map.get(
                                dept.name,
                                0
                            ) + 1
                        )

        context['overdue_department_labels'] = json.dumps(
            list(overdue_department_map.keys())
        )

        context['overdue_department_counts'] = json.dumps(
            list(overdue_department_map.values())
        )


        return context
    
# ==================================================
# AJAX VIEWS for Cascading Dropdowns
# These views must exist to support the dashboard filters.
# ==================================================

class GetZonesForPlantAjaxView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        plant_id = request.GET.get('plant_id')
        if not plant_id: return JsonResponse([], safe=False)
        zones = Zone.objects.filter(plant_id=plant_id, is_active=True).values('id', 'name')
        return JsonResponse(list(zones), safe=False)

class GetLocationsForZoneAjaxView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        zone_id = request.GET.get('zone_id')
        if not zone_id: return JsonResponse([], safe=False)
        locations = Location.objects.filter(zone_id=zone_id, is_active=True).values('id', 'name')
        return JsonResponse(list(locations), safe=False)

# This view was missing from your urls.py but is needed for the functionality
class GetSubLocationsForLocationAjaxView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        location_id = request.GET.get('location_id')
        if not location_id: return JsonResponse([], safe=False)
        sublocations = SubLocation.objects.filter(location_id=location_id, is_active=True).values('id', 'name')
        return JsonResponse(list(sublocations), safe=False)


class GetSubLocationsForLocationAjaxView(LoginRequiredMixin, TemplateView):
    """
    AJAX view to get sublocations for a selected location.
    This is called by the JavaScript on the dashboard page.
    """
    def get(self, request, *args, **kwargs):
        # Get the 'location_id' from the GET parameters of the request.
        location_id = request.GET.get('location_id')
        
        # If no location_id is provided, return an empty JSON array.
        if not location_id: 
            return JsonResponse([], safe=False)
            
        # Filter SubLocation objects that are active and belong to the selected location.
        # .values('id', 'name') ensures we only fetch the data we need.
        sublocations = SubLocation.objects.filter(location_id=location_id, is_active=True).values('id', 'name')
        
        # Return the queryset as a JSON response.
        return JsonResponse(list(sublocations), safe=False)
class ExportHazardsView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = self.request.user

        # 1. First, establish the base queryset based on user permissions.
        # This logic now mirrors the Incident export view.
        user = request.user
        if user.is_superuser or user.is_staff or getattr(user, 'is_admin_user', False):
            queryset = Hazard.objects.all()
        else:
            assigned_plants = user.assigned_plants.filter(is_active=True)
            if assigned_plants.exists():
                queryset = Hazard.objects.filter(plant__in=assigned_plants)
            elif getattr(user, 'plant', None):
                queryset = Hazard.objects.filter(plant=user.plant)
            else:
                queryset = Hazard.objects.filter(reported_by=user)

        # 2. Now, apply filters from the URL.
        selected_plant = request.GET.get('plant')
        selected_zone = request.GET.get('zone')
        selected_location = request.GET.get('location')
        selected_sublocation = request.GET.get('sublocation')
        selected_severity = request.GET.get('severity')
        selected_status = request.GET.get('status')
        selected_month = request.GET.get('month')
        selected_category = request.GET.get('category')
        selected_department = request.GET.get('department')
        selected_overdue = request.GET.get('overdue')
        selected_closed = request.GET.get('closed')

        # The plant filter from the URL is ONLY applied if the user is an Admin/Superuser.
        if selected_plant and (user.is_superuser or (hasattr(user, 'role') and user.role.name == 'ADMIN')):
            queryset = queryset.filter(plant_id=selected_plant)
        
        # Apply other filters to the permission-scoped queryset.
        if selected_zone:
            queryset = queryset.filter(zone_id=selected_zone)
        if selected_location:
            queryset = queryset.filter(location_id=selected_location)
        if selected_sublocation:
            queryset = queryset.filter(sublocation_id=selected_sublocation)
        if selected_severity:
            queryset = queryset.filter(severity__iexact=selected_severity)
        if selected_category:
            queryset = queryset.filter(hazard_category=selected_category)
        if selected_department:
            queryset = queryset.filter(
                Q(assigned_to__department_id=selected_department) |
                Q(assigned_to__department__isnull=True, reported_by__department_id=selected_department)
            )
        queryset = filter_hazards_by_status(queryset, selected_status)
        if selected_overdue == 'true':
            queryset = filter_hazards_by_status(queryset, 'OVERDUE')
        if selected_closed == 'true':
            queryset = filter_hazards_by_status(queryset, 'CLOSED')
        
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                queryset = queryset.filter(incident_datetime__year=year, incident_datetime__month=month)
            except (ValueError, TypeError):
                pass
        
        # Optimize database queries.
        queryset = queryset.select_related(
            'plant',
            'zone',
            'location',
            'sublocation',
            'reported_by',
            'reported_by__department',
            'assigned_to',
            'assigned_to__department'
        )

        # --- Excel generation code starts here ---
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Hazards Report'

        # --- Define Styles ---
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # --- Headers ---
        headers = [
            'Report Number', 'Title', 'Type', 'Category', 'Severity', 'Status',
            'Incident Datetime', 'Reported By', 'Department', 'Reported Date', 'Plant', 'Zone',
            'Location', 'Sub-Location', 'Description', 'Action Deadline'
        ]
        sheet.append(headers)
        
        # Style header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # --- Data Population ---
        for hazard in queryset:
            reporter = hazard.reported_by
            reporter_name = reporter.get_full_name().strip() if reporter else ''
            if reporter and not reporter_name:
                reporter_name = reporter.username or reporter.email or 'N/A'

            assigned_user = getattr(hazard, 'assigned_to', None)
            assigned_department = getattr(getattr(assigned_user, 'department', None), 'name', '')
            reporter_department = getattr(getattr(reporter, 'department', None), 'name', 'N/A')
            reported_by_display = reporter_name if reporter_name else 'N/A'
            department_display = assigned_department or reporter_department or 'N/A'

            row_data = [
                hazard.report_number or '',
                hazard.hazard_title or '',
                hazard.get_hazard_type_display() if hazard.hazard_type else '',
                hazard.get_hazard_category_display() if hazard.hazard_category else '',
                hazard.get_severity_display() if hazard.severity else '',
                hazard.effective_status_display if hazard.status else '',
                hazard.incident_datetime.strftime('%Y-%m-%d %H:%M') if hazard.incident_datetime else '',
                reported_by_display,
                department_display,
                hazard.created_at.strftime('%Y-%m-%d') if hazard.created_at else '',
                hazard.plant.name if hazard.plant else 'N/A',
                hazard.zone.name if hazard.zone else 'N/A',
                hazard.location.name if hazard.location else 'N/A',
                hazard.sublocation.name if hazard.sublocation else 'N/A',
                hazard.hazard_description or '',
                hazard.action_deadline.strftime('%Y-%m-%d') if hazard.action_deadline else ''
            ]
            sheet.append(row_data)

        # --- Auto-adjust Column Widths and Apply Wrapping ---
        desc_col_letter = get_column_letter(headers.index('Description') + 1)
        title_col_letter = get_column_letter(headers.index('Title') + 1)

        for col_idx, column_cells in enumerate(sheet.columns, 1):
            column_letter = get_column_letter(col_idx)
            # Set a fixed width for columns that need text wrapping
            if column_letter in [desc_col_letter, title_col_letter]:
                sheet.column_dimensions[column_letter].width = 50
                # Apply wrap text to all cells in the description/title column
                for cell in column_cells:
                    cell.alignment = wrap_alignment
            else:
                # Auto-size other columns
                max_length = 0
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2

        # --- HTTP Response ---
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"Hazards_Report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response
    
    
    

class HazardPDFView(LoginRequiredMixin, View):
    """
    Handles the generation and download of a Hazard report in PDF format.
    """
    def get(self, request, *args, **kwargs):
        """
        Processes the GET request to download the PDF for a specific hazard.
        """
        # Retrieve the primary key of the hazard from the URL.
        hazard_pk = self.kwargs.get('pk')

        # Fetch the Hazard object from the database, or return a 404 error if not found.
        # This pre-fetches related objects to optimize database queries.
        hazard = get_object_or_404(
            Hazard.objects.select_related(
                'plant', 'zone', 'location', 'sublocation', 
                'reported_by', 'behalf_person_dept'
            ), 
            pk=hazard_pk
        )
        # Permission check
        if not (
            request.user.is_superuser or request.user == hazard.reported_by or
            request.user.has_permission('EXPORT_HAZARD_PDF')):
            messages.error(request, "You don't have permission to view this report")
            return redirect('hazards:hazard_list')
                        
        # Call the PDF generation utility function and return its response.
        return generate_hazard_pdf(hazard)
    
    
class HazardApprovalView(LoginRequiredMixin, DetailView):
    """
    Displays hazard summary for approval and handles the POST requests
    for approving or rejecting the hazard report.
    """
    model = Hazard
    template_name = 'hazards/hazard_approval.html'
    context_object_name = 'hazard'

    def get_context_data(self, **kwargs):
        """Adds related action items to the context."""
        context = super().get_context_data(**kwargs)
        # Fetch all action items related to this hazard to display them.
        context['action_items'] = self.object.action_items.all()
        return context

    def post(self, request, *args, **kwargs):
        """Handles the 'approve' and 'reject' form submissions."""
        hazard = self.get_object()

        # Check which button was clicked based on its 'name' attribute in the form.
        if 'approve_action' in request.POST:
            # Handle the approval logic.
            hazard.approval_status = 'APPROVED'
            hazard.approved_by = request.user
            hazard.approved_date = timezone.now()
            hazard.approved_remarks = ""  # Clear any previous remarks.
            
            # ✅ CRITICAL FIX: Check if action items exist
            action_items_exist = hazard.action_items.exists()
            
            if action_items_exist:
                # If action items exist, set status to ACTION_ASSIGNED
                hazard.status = 'ACTION_ASSIGNED'
            else:
                # If no action items, set to APPROVED
                hazard.status = 'APPROVED'
            
            hazard.save()
            
            # ✅ NOW call update to ensure correct status based on action item progress
            # This will only work because status is no longer PENDING_APPROVAL
            hazard.update_status_from_action_items()

            messages.success(request, f"Hazard {hazard.report_number} has been approved and is now active.")

        elif 'reject_action' in request.POST:
            # Handle the rejection logic.
            rejection_remarks = request.POST.get('rejection_remarks', '').strip()
            if not rejection_remarks:
                messages.error(request, "Rejection remarks are required to reject the report.")
                return redirect('hazards:hazard_approve', pk=hazard.pk)

            hazard.status = 'REJECTED'
            hazard.approval_status = 'REJECTED'
            hazard.approved_remarks = rejection_remarks
            hazard.approved_by = None
            hazard.approved_date = None
            hazard.save()
            messages.warning(request, f"Hazard {hazard.report_number} has been rejected.")

        return redirect('hazards:hazard_detail', pk=hazard.pk)
    

# Add these RIGHT AFTER your imports, BEFORE the class-based views

from django.views.decorators.http import require_GET

@require_GET
def get_zones_by_plant(request, plant_id):
    """
    Function-based view for cascading dropdown
    URL: /hazards/api/get-zones/<plant_id>/
    """
    try:
        user = request.user
        if not user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        if not user.assigned_plants.filter(id=plant_id, is_active=True).exists():
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        zones = Zone.objects.filter(
            plant_id=plant_id, 
            is_active=True
        ).values('id', 'name', 'code').order_by('sequence')
        return JsonResponse(list(zones), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def get_locations_by_zone(request, zone_id):
    """
    Function-based view for cascading dropdown
    URL: /hazards/api/get-locations/<zone_id>/
    """
    try:
        user = request.user
        if not user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        # Check if user has access to the plant that contains this zone
        if not user.assigned_plants.filter(
            zones__id=zone_id, 
            is_active=True
        ).exists():
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        locations = Location.objects.filter(
            zone_id=zone_id, 
            is_active=True
        ).values('id', 'name', 'code').order_by('name')
        return JsonResponse(list(locations), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def get_sublocations_by_location(request, location_id):
    """
    Function-based view for cascading dropdown
    URL: /hazards/api/get-sublocations/<location_id>/
    """
    try:
        user = request.user
        if not user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        # Check if user has access to the plant that contains this location's zone
        if not user.assigned_plants.filter(
            zones__locations__id=location_id,
            is_active=True
        ).exists():
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        sublocations = SubLocation.objects.filter(
            location_id=location_id,
            is_active=True
        ).values('id', 'name', 'code').order_by('name')
        return JsonResponse(list(sublocations), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)    
    

class MyActionItemsView(LoginRequiredMixin, ListView):
    """
    Display action items assigned to the logged-in user, reflecting their
    personal completion status.
    """
    model = HazardActionItem
    template_name = 'hazards/my_action_items.html'
    context_object_name = 'action_items'
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user
        
        # Base queryset for items assigned to the user
        queryset = HazardActionItem.objects.filter(
            responsible_emails__icontains=user.email
        ).select_related(
            'hazard', 
            'hazard__plant', 
            'hazard__location',
            'created_by'
        ).prefetch_related(
            'completed_by_users'  # <-- CRITICAL: Prefetch for efficiency
        ).order_by('-created_at')
        
        # Apply filters
        status_filter = self.request.GET.get('status', '')
        severity_filter = self.request.GET.get('severity', '')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if severity_filter:
            queryset = queryset.filter(hazard__severity=severity_filter)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        user = self.request.user
        
        # Get the full (unpaginated) queryset for accurate stats
        all_my_items = self.get_queryset()
        
        # --- USER-SPECIFIC STATISTICS ---
        
        # Items the logged-in user has personally completed
        my_completed_items = all_my_items.filter(completed_by_users=user)
        
        # Items assigned to the user that they have NOT yet completed
        my_pending_items = all_my_items.exclude(completed_by_users=user)

        context['total_assigned'] = all_my_items.count()
        context['completed_count'] = my_completed_items.count()
        context['pending_count'] = my_pending_items.count()
        
        # An item is "In Progress" if it's pending for the current user,
        # but has been completed by at least one other person.
        context['in_progress_count'] = my_pending_items.filter(
            status='IN_PROGRESS'
        ).count()
        
        # Overdue items are those pending for the current user that are past their target date.
        context['overdue_count'] = my_pending_items.filter(
            target_date__lt=timezone.now().date()
        ).count()
        
        # --- ANNOTATE EACH ITEM WITH USER'S COMPLETION STATUS ---
        # Get the list of items for the current page
        action_items_on_page = context['action_items']
        
        for item in action_items_on_page:
            # Check if the current user is among those who completed this item.
            # This is efficient because of the prefetch_related in get_queryset.
            completed_user_ids = {u.id for u in item.completed_by_users.all()}
            if user.id in completed_user_ids:
                item.user_has_completed = True
            else:
                item.user_has_completed = False
        
        # Filter values
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_severity'] = self.request.GET.get('severity', '')
        
        # Choices for filters
        context['status_choices'] = HazardActionItem.STATUS_CHOICES
        context['severity_choices'] = Hazard.SEVERITY_CHOICES
        
        return context
    

class ActionItemCompleteView(LoginRequiredMixin, UpdateView):
    """
    Allow assigned users to mark action item as complete
    """
    model = HazardActionItem
    template_name = 'hazards/action_item_complete.html'
    fields = []
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(f"{reverse('accounts:login')}?next={request.path}")
        self.object = self.get_object()
        
        user_email = request.user.email
        if user_email not in self.object.responsible_emails:
            messages.error(request, 'You are not assigned to this action item.')
            return redirect('hazards:my_action_items')

        if self.object.status == 'COMPLETED':
            messages.info(request, 'This action item is already completed.')
            return redirect('hazards:my_action_items')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_item'] = self.object
        context['hazard'] = self.object.hazard
        return context
    
    def post(self, request, *args, **kwargs):
        action_item = self.get_object()
        
        try:
            completion_remarks = request.POST.get('completion_remarks', '').strip()
            
            if not completion_remarks:
                messages.error(request, 'Completion remarks are required.')
                return redirect('hazards:action_item_complete', pk=action_item.pk)

            # --- MODIFIED LOGIC ---
            
            # 1. Add the current user to the list of users who have completed the action.
            action_item.completed_by_users.add(request.user)
            
            # 2. Update remarks and attachment. The last person's remarks will be saved.
            action_item.completion_remarks = completion_remarks
            if 'completion_attachment' in request.FILES:
                action_item.attachment = request.FILES['completion_attachment']
            
            # 3. Save the action item. The model's save() method will now automatically handle
            # updating the status to 'IN_PROGRESS' or 'COMPLETED'.
            action_item.save()
            
            # 4. Trigger the parent hazard's status update.
            action_item.hazard.update_status_from_action_items()
            
            # --- END OF MODIFIED LOGIC ---

            try:
                from apps.notifications.services import NotificationService
                NotificationService.notify(
                    content_object=action_item,
                    notification_type='HAZARD_ACTION_COMPLETED',
                    module='HAZARD_ACTION'
                )
            except Exception as e:
                print(f"Notification error: {e}")
            
            messages.success(
                request,
                mark_safe(
                    f'✅ <strong>Action item marked as completed!</strong><br>'
                    f'Hazard: {action_item.hazard.report_number}'
                )
            )
            
            return redirect('hazards:my_action_items')
            
        except Exception as e:
            print(f"Error completing action item: {e}")
            messages.error(request, f'Error: {str(e)}')
            return redirect('hazards:action_item_complete', pk=action_item.pk)
