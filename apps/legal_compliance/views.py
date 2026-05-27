from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.db.models import Count
from django.utils import timezone
from datetime import timedelta
from django.urls import reverse_lazy
from django.views.generic import (TemplateView, ListView, CreateView, UpdateView,DeleteView, DetailView,)
from .models import (ComplianceQuestion, ComplianceRequirementQuestion,LegalAct,ComplianceRequirement,ComplianceSubmission,ComplianceResponse, ComplianceFinding)
from .forms import (ComplianceQuestionFilterForm, ComplianceQuestionForm, LegalActForm, ComplianceRequirementForm, User, ComplianceRequirementForm )
from django.shortcuts import (get_object_or_404,redirect, render)
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from reportlab.platypus import (SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from io import BytesIO
from .mixins import (LegalComplianceViewMixin,LegalComplianceConfigMixin)
from django.core.paginator import Paginator
from django.db import transaction
from reportlab.platypus import (SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer)
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus.flowables import PageBreak

# =========================================================
# DASHBOARD
# =========================================================

class LegalComplianceDashboardView(LoginRequiredMixin, TemplateView):

    template_name = 'legal_compliance/dashboard.html'

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context['total_acts'] = LegalAct.objects.count()

        context['active_acts'] = LegalAct.objects.filter(
            is_active=True
        ).count()

        context['total_compliances'] = ComplianceRequirement.objects.count()

        context['active_compliances'] = ComplianceRequirement.objects.filter(
            is_active=True
        ).count()

        context['critical_compliances'] = ComplianceRequirement.objects.filter(
            criticality='CRITICAL'
        ).count()

        context['recent_acts'] = LegalAct.objects.order_by(
            '-created_at'
        )[:5]

        context['recent_compliances'] = ComplianceRequirement.objects.select_related(
            'legal_act'
        ).order_by('-created_at')[:5]

        return context


# =========================================================
# LEGAL ACT VIEWS
# =========================================================

class LegalActListView(LoginRequiredMixin,LegalComplianceViewMixin,ListView):

    model = LegalAct

    template_name = 'legal_compliance/acts/act_list.html'

    context_object_name = 'acts'

    paginate_by = 10

    def get_queryset(self):

        queryset = LegalAct.objects.all()

        search = self.request.GET.get('search')

        category = self.request.GET.get('category')

        government_level = self.request.GET.get('government_level')

        if search:
            queryset = queryset.filter(
                act_name__icontains=search
            )

        if category:
            queryset = queryset.filter(
                category=category
            )

        if government_level:
            queryset = queryset.filter(
                government_level=government_level
            )

        return queryset.order_by('act_name')

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context['categories'] = LegalAct.CATEGORY_CHOICES

        context['government_levels'] = LegalAct.GOVERNMENT_LEVEL_CHOICES

        return context


class LegalActCreateView(LoginRequiredMixin,LegalComplianceConfigMixin,CreateView):

    model = LegalAct

    form_class = LegalActForm

    template_name = 'legal_compliance/acts/act_form.html'

    success_url = reverse_lazy(
        'legal_compliance:act_list'
    )

    def form_valid(self, form):

        form.instance.created_by = self.request.user

        messages.success(
            self.request,
            'Legal Act created successfully.'
        )

        return super().form_valid(form)


class LegalActDetailView(LoginRequiredMixin,LegalComplianceViewMixin,DetailView):

    model = LegalAct

    template_name = 'legal_compliance/acts/act_detail.html'

    context_object_name = 'act'

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context['compliances'] = self.object.requirements.all()

        return context


class LegalActUpdateView(LoginRequiredMixin,LegalComplianceConfigMixin,UpdateView):

    model = LegalAct

    form_class = LegalActForm

    template_name = 'legal_compliance/acts/act_form.html'

    success_url = reverse_lazy(
        'legal_compliance:act_list'
    )

    def form_valid(self, form):

        messages.success(
            self.request,
            'Legal Act updated successfully.'
        )

        return super().form_valid(form)


class LegalActDeleteView(LoginRequiredMixin,LegalComplianceConfigMixin,DeleteView):

    model = LegalAct

    template_name = 'legal_compliance/acts/act_confirm_delete.html'

    success_url = reverse_lazy(
        'legal_compliance:act_list'
    )

    def delete(self, request, *args, **kwargs):

        messages.success(
            self.request,
            'Legal Act deleted successfully.'
        )

        return super().delete(request, *args, **kwargs)


# =========================================================
# COMPLIANCE Questions
# =========================================================
# =========================================================
# COMPLIANCE QUESTION VIEWS
# =========================================================

@login_required
def compliance_question_list(request):
    """
    List all compliance questions with filters
    """

    questions = (

        ComplianceQuestion.objects
        .select_related('legal_act')
        .filter(is_active=True)
    )

    # FILTERS

    filter_form = ComplianceQuestionFilterForm(
        request.GET
    )

    if filter_form.is_valid():

        legal_act = (
            filter_form.cleaned_data.get(
                'legal_act'
            )
        )

        submission_type = (
            filter_form.cleaned_data.get(
                'submission_type'
            )
        )

        is_critical = (
            filter_form.cleaned_data.get(
                'is_critical'
            )
        )

        search = (
            filter_form.cleaned_data.get(
                'search'
            )
        )

        if legal_act:

            questions = questions.filter(
                legal_act=legal_act
            )

        if submission_type:

            questions = questions.filter(
                submission_type=submission_type
            )

        if is_critical is not None:

            questions = questions.filter(
                is_critical=is_critical
            )

        if search:

            questions = questions.filter(

                Q(question_text__icontains=search)

                |

                Q(question_code__icontains=search)

                |

                Q(reference_standard__icontains=search)
            )

    questions = questions.order_by(
        'legal_act'
    )

    # PAGINATION

    paginator = Paginator(
        questions,
        25
    )

    page_number = request.GET.get(
        'page'
    )

    page_obj = paginator.get_page(
        page_number
    )

    context = {

        'page_obj': page_obj,

        'filter_form': filter_form,

        'total_questions': questions.count()
    }

    return render(

        request,

        'legal_compliance/questions/question_list.html',

        context
    )



@login_required
def compliance_question_create(request):
    """
    Create compliance question
    """

    if request.method == 'POST':

        form = ComplianceQuestionForm(
            request.POST
        )

        if form.is_valid():

            question = form.save(
                commit=False
            )

            question.created_by = (
                request.user
            )

            question.save()

            form.save_m2m()

            messages.success(

                request,

                f'Question "{question.question_code}" created successfully!'
            )

            # SAVE & ADD ANOTHER

            if (
                request.POST.get(
                    'action_type'
                ) == 'save_and_add'
            ):

                return redirect(
                    'legal_compliance:question_create'
                )

            return redirect(
                'legal_compliance:question_list'
            )

    else:

        form = ComplianceQuestionForm()

        # PRESELECT LEGAL ACT

        legal_act_id = request.GET.get(
            'legal_act'
        )

        if legal_act_id:

            form.initial[
                'legal_act'
            ] = legal_act_id

    context = {

        'form': form,

        'action': 'Create',

        'title': 'Create Compliance Question'
    }

    return render(

        request,

        'legal_compliance/questions/question_form.html',

        context
    )



@login_required
def compliance_question_edit(request, pk):
    """
    Edit compliance question
    """

    question = get_object_or_404(
        ComplianceQuestion,
        pk=pk
    )

    if request.method == 'POST':

        form = ComplianceQuestionForm(

            request.POST,

            instance=question
        )

        if form.is_valid():

            question = form.save(
                commit=False
            )

            question.updated_by = (
                request.user
            )

            question.save()

            form.save_m2m()

            messages.success(

                request,

                f'Question "{question.question_code}" updated successfully!'
            )

            return redirect(
                'legal_compliance:question_list'
            )

    else:

        form = ComplianceQuestionForm(
            instance=question
        )

    context = {

        'form': form,

        'action': 'Edit',

        'title': (
            f'Edit Question: '
            f'{question.question_code}'
        ),

        'question': question
    }

    return render(

        request,

        'legal_compliance/questions/question_form.html',

        context
    )


@login_required
def compliance_question_detail(request, pk):
    """
    View compliance question details
    """

    question = get_object_or_404(

        ComplianceQuestion.objects.select_related(

            'legal_act',

            'created_by',

            'updated_by'
        ).prefetch_related(

            'applicable_plants',

            'applicable_departments'
        ),

        pk=pk
    )

    # GET REQUIREMENTS USING THIS QUESTION

    requirements = (

        ComplianceRequirement.objects.filter(

            requirement_questions__question=question,

            is_active=True
        )

        .distinct()

        .order_by('title')
    )

    context = {

        'question': question,

        'requirements': requirements,

        'requirements_count': requirements.count()
    }

    return render(

        request,

        'legal_compliance/questions/question_detail.html',

        context
    )

@login_required
def compliance_question_delete(request, pk):
    """
    Soft delete compliance question
    """

    question = get_object_or_404(
        ComplianceQuestion,
        pk=pk
    )

    if request.method == 'POST':

        question.is_active = False

        question.save()

        messages.success(

            request,

            f'Question "{question.question_code}" deleted successfully!'
        )

        return redirect(
            'legal_compliance:question_list'
        )

    # REQUIREMENT COUNT

    requirements_count = (

        ComplianceRequirement.objects.filter(

            requirement_questions__question=question

        )

        .distinct()

        .count()
    )

    context = {

        'question': question,

        'requirements_count': requirements_count
    }

    return render(

        request,

        'legal_compliance/questions/question_confirm_delete.html',

        context
    )





# =========================================================
# COMPLIANCE compliance VIEWS
# =========================================================

class ComplianceRequirementListView(LoginRequiredMixin, ListView):

    model = ComplianceRequirement

    template_name = 'legal_compliance/compliance/compliance_list.html'

    context_object_name = 'compliances'

    paginate_by = 10

    def get_queryset(self):

        queryset = (
            ComplianceRequirement.objects
            .select_related(
                'legal_act'
            )
            .prefetch_related(
                'responsible_person',
                'reviewer'
            )
        )

        search = self.request.GET.get('search')

        frequency = self.request.GET.get('frequency')

        criticality = self.request.GET.get('criticality')

        if search:
            queryset = queryset.filter(
                title__icontains=search
            )

        if frequency:
            queryset = queryset.filter(
                frequency=frequency
            )

        if criticality:
            queryset = queryset.filter(
                criticality=criticality
            )

        return queryset.order_by('title')

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context['frequencies'] = ComplianceRequirement.FREQUENCY_CHOICES

        context['criticalities'] = ComplianceRequirement.CRITICALITY_CHOICES

        return context
    
from django.shortcuts import redirect
from django.contrib import messages


class ComplianceRequirementCreateView(
    LoginRequiredMixin,
    CreateView
):

    model = ComplianceRequirement

    form_class = ComplianceRequirementForm

    template_name = (
        'legal_compliance/compliance/compliance_form.html'
    )

    success_url = reverse_lazy(
        'legal_compliance:compliance_list'
    )

    def form_valid(self, form):

        requirement = form.save(commit=False)

        requirement.created_by = self.request.user

        requirement.save()

        # ============================================
        # SAVE M2M FROM FORM
        # ============================================

        form.save_m2m()

        # ============================================
        # SAVE RESPONSIBLE PERSONS
        # ============================================

        responsible_users = self.request.POST.getlist(
            'responsible_person'
        )

        if responsible_users:

            requirement.responsible_person.set(
                responsible_users
            )

        # ============================================
        # SAVE REVIEWERS
        # ============================================

        reviewers = self.request.POST.getlist(
            'reviewer'
        )

        if reviewers:

            requirement.reviewer.set(
                reviewers
            )

        # ============================================
        # AUTO LINK QUESTIONS
        # ============================================

        questions = ComplianceQuestion.objects.filter(

            legal_act=requirement.legal_act,

            is_active=True
        )

        for question in questions:

            ComplianceRequirementQuestion.objects.get_or_create(

                compliance_requirement=requirement,

                question=question
            )

        messages.success(

            self.request,

            'Compliance Requirement created successfully.'
        )

        return redirect(
            self.success_url
        ) 


class ComplianceRequirementDetailView(LoginRequiredMixin, DetailView):

    model = ComplianceRequirement

    template_name = 'legal_compliance/compliance/compliance_detail.html'

    context_object_name = 'compliance'


class ComplianceRequirementUpdateView(
    LoginRequiredMixin,
    UpdateView
):

    model = ComplianceRequirement

    form_class = ComplianceRequirementForm

    template_name = (
        'legal_compliance/compliance/compliance_form.html'
    )

    success_url = reverse_lazy(
        'legal_compliance:compliance_list'
    )

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        compliance = self.object

        context['selected_plant_ids'] = list(

            compliance.applicable_plants.values_list(
                'id',
                flat=True
            )
        )

        context['selected_responsible_ids'] = list(

            compliance.responsible_person.values_list(
                'id',
                flat=True
            )
        )

        context['selected_reviewer_ids'] = list(

            compliance.reviewer.values_list(
                'id',
                flat=True
            )
        )

        return context

    def form_valid(self, form):

        requirement = form.save()

        responsible_users = self.request.POST.getlist(
            'responsible_person'
        )

        reviewers = self.request.POST.getlist(
            'reviewer'
        )

        requirement.responsible_person.set(
            responsible_users
        )

        requirement.reviewer.set(
            reviewers
        )

        messages.success(

            self.request,

            'Compliance updated successfully.'
        )

        return redirect(
            self.success_url
        )


class ComplianceRequirementDeleteView(LoginRequiredMixin, DeleteView):

    model = ComplianceRequirement

    template_name = 'legal_compliance/compliance/compliance_confirm_delete.html'

    success_url = reverse_lazy(
        'legal_compliance:compliance_list'
    )

    def delete(self, request, *args, **kwargs):

        messages.success(
            self.request,
            'Compliance compliance deleted successfully.'
        )

        return super().delete(request, *args, **kwargs)
    


# =====================================================
# Fetch User from Plant
# =====================================================
@login_required
def get_users_by_plants(request):
    """
    AJAX: Get HODs and Safety Managers for selected plants.
    """

    plant_ids = request.GET.get('plant_ids', '')

    if not plant_ids:

        return JsonResponse({
            'users': []
        })

    ids = [

        pid.strip()

        for pid in plant_ids.split(',')

        if pid.strip()
    ]

    users = (

        User.objects.filter(

            plant__id__in=ids,

            role__name__in=[
                'HOD',
                'SAFETY MANAGER'
            ],

            is_active_employee=True,

            is_active=True

        )

        .select_related(
            'plant',
            'role',
            'department'
        )

        .order_by(
            'plant__name',
            'first_name'
        )
    )

    users_data = []

    for u in users:

        users_data.append({

            'id': u.id,

            'full_name': u.get_full_name(),

            'role': (
                u.role.name
                if u.role else ''
            ),

            'department': (
                u.department.name
                if u.department else ''
            ),

            'plant_name': (
                u.plant.name
                if u.plant else ''
            ),

            'plant_id': (
                u.plant.id
                if u.plant else None
            ),
        })

    return JsonResponse({
        'users': users_data
    })



@login_required
def get_reviewers(request):

    plant_ids = request.GET.get('plant_ids', '')

    if not plant_ids:

        return JsonResponse({
            'users': []
        })

    ids = [

        pid.strip()

        for pid in plant_ids.split(',')

        if pid.strip()
    ]

    users = (

        User.objects.filter(

            plant__id__in=ids,

            department_id=2,
            role__name__in=[
                'ADMIN',
                'HOD',
                'SAFETY MANAGER'
            ],

            is_active_employee=True,

            is_active=True

        )

        .select_related(
            'department',
            'role',
            'plant'
        )

        .order_by(
            'plant__name',
            'first_name'
        )
    )

    users_data = []

    for u in users:

        users_data.append({

            'id': u.id,

            'full_name': u.get_full_name(),

            'role': (
                u.role.name
                if u.role else ''
            ),

            'department': (
                u.department.name
                if u.department else ''
            ),

            'plant_name': (
                u.plant.name
                if u.plant else ''
            ),

            'plant_id': (
                u.plant.id
                if u.plant else None
            ),
        })

    return JsonResponse({
        'users': users_data
    })


# =====================================================
# MY COMPLIANCES
# =====================================================
@login_required
def my_compliances(request):

    # =====================================================
    # RESPONSIBLE PERSON COMPLIANCES
    # =====================================================

    responsible_queryset = (

        ComplianceRequirement.objects

        .filter(

            responsible_person=request.user,

            is_active=True
        )
    )

    # =====================================================
    # REVIEWER COMPLIANCES
    # ONLY SUBMITTED / COMPLETED
    # =====================================================

    reviewer_queryset = (

        ComplianceRequirement.objects

        .filter(

            reviewer=request.user,

            is_active=True,

            status__in=[
                'SUBMITTED',
                'COMPLETED'
            ]
        )
    )

    # =====================================================
    # FINAL QUERYSET
    # =====================================================

    compliances = (

        responsible_queryset

        |

        reviewer_queryset
    ).distinct()

    compliances = (

        compliances

        .select_related(
            'legal_act'
        )

        .prefetch_related(
            'requirement_questions__question',
            'responsible_person',
            'reviewer'
        )
    )

    # =====================================================
    # STATS
    # =====================================================

    stats = {

        'total': compliances.count(),

        'pending': compliances.filter(
            status='PENDING'
        ).count(),

        'in_progress': compliances.filter(
            status='IN_PROGRESS'
        ).count(),

        'submitted': compliances.filter(
            status='SUBMITTED'
        ).count(),

        'overdue': compliances.filter(
            status='OVERDUE'
        ).count(),
    }

    # =====================================================
    # FILTERS
    # =====================================================

    status = request.GET.get(
        'status'
    )

    if status:

        compliances = compliances.filter(
            status=status
        )

    # =====================================================
    # ORDERING
    # =====================================================

    compliances = compliances.order_by(
        '-scheduled_date',
        '-created_at'
    )

    # =====================================================
    # PAGINATION
    # =====================================================

    paginator = Paginator(
        compliances,
        15
    )

    page_number = request.GET.get(
        'page'
    )

    page_obj = paginator.get_page(
        page_number
    )

    context = {

        'page_obj': page_obj,

        'stats': stats,

        'selected_status': status,

        'status_choices': ComplianceRequirement.STATUS_CHOICES
    }

    return render(

        request,

        'legal_compliance/my_compliances.html',

        context
    )

# =====================================================
# START COMPLIANCE
# =====================================================

@login_required
def compliance_start(request, requirement_id):

    requirement = get_object_or_404(

        ComplianceRequirement.objects.prefetch_related(
            'questions'
        ),

        pk=requirement_id
    )

    if requirement.status == 'PENDING':

        requirement.status = 'IN_PROGRESS'

        requirement.started_at = timezone.now()

        requirement.save()

    if request.user not in requirement.responsible_person.all():

        messages.error(
            request,
            'Unauthorized access.'
        )

        return redirect(
            'legal_compliance:my_compliances'
        )
    
    questions = (

        ComplianceRequirementQuestion.objects

        .filter(
            compliance_requirement=requirement
        )

        .select_related(
            'question'
        )

        .order_by('id')
    )

    # ==========================================
    # GET LAST SUBMISSION
    # ==========================================

    last_submission = (

        ComplianceSubmission.objects

        .filter(
            requirement=requirement
        )

        .prefetch_related(
            'responses'
        )

        .order_by('-created_at')

        .first()
    )

    existing_responses = {}

    reviewer_comments = None

    if last_submission:

        reviewer_comments = (
            last_submission.reviewer_comments
        )

        for response in last_submission.responses.all():

            existing_responses[
                response.question_id
            ] = response


    context = {

        'requirement': requirement,

        'questions': questions,

        'last_submission': last_submission,

        'existing_responses': existing_responses,

        'reviewer_comments': reviewer_comments,
    }

    return render(

        request,

        'legal_compliance/compliance_start.html',

        context
    )


# =====================================================
# SUBMIT COMPLIANCE
# =====================================================

@login_required
@transaction.atomic
def compliance_submit(request, requirement_id):

    requirement = get_object_or_404(

        ComplianceRequirement,

        pk=requirement_id
    )

    if request.method != 'POST':

        return redirect(

            'legal_compliance:compliance_start',

            requirement_id=requirement.id
        )

    submission = (

        ComplianceSubmission.objects.create(

            requirement=requirement,

            submitted_by=request.user,

            status='SUBMITTED',

            remarks=request.POST.get(
                'overall_remarks',
                ''
            ),

            submitted_at=timezone.now()
        )
    )

    questions = (

        ComplianceRequirementQuestion.objects

        .filter(
            compliance_requirement=requirement
        )

        .select_related(
            'question'
        )
    )

    for mapping in questions:

        question = mapping.question

        ComplianceResponse.objects.create(

            submission=submission,

            question=question,

            answer=request.POST.get(
                f'answer_{question.id}'
            ),

            remarks=request.POST.get(
                f'remarks_{question.id}'
            ),

            evidence_file=request.FILES.get(
                f'evidence_{question.id}'
            )
        )

    # ==========================================
    # UPDATE REQUIREMENT STATUS
    # ==========================================

    requirement.status = 'SUBMITTED'
    requirement.save()

    messages.success(
        request,
        'Compliance submitted successfully.'
    )

    return redirect(
        'legal_compliance:my_compliances'
    )


# =====================================================
# REVIEW COMPLIANCE
# =====================================================
@login_required
def compliance_review(request, submission_id):

    submission = get_object_or_404(

        ComplianceSubmission.objects.select_related(

            'requirement',

            'submitted_by'
        ),

        pk=submission_id
    )

    responses = (

        submission.responses.select_related(
            'question'
        )
    )

    # ==========================================
    # COUNTS
    # ==========================================

    total_questions = responses.count()

    yes_count = responses.filter(
        answer='YES'
    ).count()

    no_count = responses.filter(
        answer='NO'
    ).count()

    na_count = responses.filter(
        answer='NA'
    ).count()

    # ==========================================
    # REVIEW ACTION
    # ==========================================

    if request.method == 'POST':

        action = request.POST.get(
            'action'
        )

        reviewer_comments = request.POST.get(
            'reviewer_comments'
        )

        if action == 'approve':

            submission.status = 'APPROVED'

            submission.reviewed_by = request.user

            submission.reviewed_at = timezone.now()

            submission.reviewer_comments = (reviewer_comments)
            submission.save()
            # ==========================================
            # MARK COMPLETED
            # ==========================================
            submission.requirement.status = 'COMPLETED'
            submission.requirement.completed_at = timezone.now()
            submission.requirement.save()



            messages.success(
                request,
                'Compliance approved successfully.'
            )

        elif action == 'reject':

            submission.status = 'REJECTED'

            submission.reviewed_by = request.user

            submission.reviewed_at = timezone.now()

            submission.reviewer_comments = (reviewer_comments)

            submission.save()

            # ==========================================
            # CREATE COMPLIANCE FINDING
            # ==========================================

            ComplianceFinding.objects.create(

                requirement=submission.requirement,

                submission=submission,

                finding_description=(reviewer_comments),

                violated_provision=(submission.requirement.legal_act.act_name),

                corrective_action='',

                responsible_person=(submission.submitted_by),

                reviewer=request.user,

                target_date=(timezone.now().date() + timedelta(days=7)),

                status='OPEN',

                created_by=request.user
            )

            submission.requirement.status = (
                'REJECTED'
            )

            submission.requirement.save()

            messages.warning(
                request,
                'Compliance rejected.'
            )

        return redirect(
            'legal_compliance:compliance_review',
            submission_id=submission.id
        )

    context = {

        'submission': submission,

        'responses': responses,

        'total_questions': total_questions,

        'yes_count': yes_count,

        'no_count': no_count,

        'na_count': na_count
    }

    return render(

        request,

        'legal_compliance/compliance_review.html',

        context
    )




# =====================================================
# COMPLIANCE STATUS REPORT
# =====================================================

class ComplianceStatusReportView(
    LoginRequiredMixin,
    TemplateView
):

    template_name = (
        'legal_compliance/reports/compliance_status_report.html'
    )

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        submissions = (

            ComplianceSubmission.objects

            .select_related(
                'requirement',
                'requirement__legal_act',
                'reviewed_by',
                'submitted_by'
            )

            .prefetch_related(
                'requirement__responsible_person',
                'requirement__reviewer',
                'requirement__applicable_plants'
            )
        )

        # =================================================
        # FILTERS
        # =================================================

        status = self.request.GET.get('status')

        legal_act = self.request.GET.get('legal_act')

        frequency = self.request.GET.get('frequency')

        if status:

            submissions = submissions.filter(
                requirement__status=status
            )

        if legal_act:

            submissions = submissions.filter(
                requirement__legal_act_id=legal_act
            )

        if frequency:

            submissions = submissions.filter(
                requirement__frequency=frequency
            )

        # =================================================
        # STATS
        # =================================================

        context['stats'] = {

            'total': submissions.count(),

            'pending': submissions.filter(
                requirement__status='PENDING'
            ).count(),

            'in_progress': submissions.filter(
                requirement__status='IN_PROGRESS'
            ).count(),

            'submitted': submissions.filter(
                requirement__status='SUBMITTED'
            ).count(),

            'completed': submissions.filter(
                requirement__status='COMPLETED'
            ).count(),

            'overdue': submissions.filter(
                requirement__status='OVERDUE'
            ).count(),

            'rejected': submissions.filter(
                requirement__status='REJECTED'
            ).count(),
        }

        context['submissions'] = submissions.order_by(
            '-created_at'
        )

        context['legal_acts'] = LegalAct.objects.filter(
            is_active=True
        )

        context['status_choices'] = (
            ComplianceRequirement.STATUS_CHOICES
        )

        context['frequency_choices'] = (
            ComplianceRequirement.FREQUENCY_CHOICES
        )

        return context
    


# =====================================================
# EXPORT COMPLIANCE STATUS REPORT EXCEL
# =====================================================

@login_required
def export_compliance_status_excel(request):

    submissions = (

        ComplianceSubmission.objects

        .select_related(
            'requirement',
            'requirement__legal_act',
            'submitted_by',
            'reviewed_by'
        )

        .prefetch_related(
            'requirement__responsible_person',
            'requirement__reviewer'
        )

        .order_by('-created_at')
    )

    # =================================================
    # CREATE WORKBOOK
    # =================================================

    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    worksheet.title = (
        'Compliance Status Report'
    )

    # =================================================
    # HEADERS
    # =================================================

    headers = [

        'Compliance Code',

        'Compliance Title',

        'Legal Act',

        'Frequency',

        'Responsible Person',

        'Reviewer',

        'Due Date',

        'Status',

        'Submitted By',

        'Submitted At',

        'Reviewed By',

        'Reviewed At'
    ]

    # =================================================
    # WRITE HEADERS
    # =================================================

    for col_num, header in enumerate(headers, 1):

        cell = worksheet.cell(

            row=1,

            column=col_num
        )

        cell.value = header

        cell.font = Font(
            bold=True
        )

    # =================================================
    # WRITE DATA
    # =================================================

    row_num = 2

    for submission in submissions:

        responsible_persons = ", ".join([

            user.get_full_name()

            for user in submission.requirement.responsible_person.all()
        ])

        reviewers = ", ".join([

            user.get_full_name()

            for user in submission.requirement.reviewer.all()
        ])

        worksheet.cell(
            row=row_num,
            column=1
        ).value = (
            submission.requirement.requirement_code
        )

        worksheet.cell(
            row=row_num,
            column=2
        ).value = (
            submission.requirement.title
        )

        worksheet.cell(
            row=row_num,
            column=3
        ).value = (
            submission.requirement.legal_act.short_name
        )

        worksheet.cell(
            row=row_num,
            column=4
        ).value = (
            submission.requirement.get_frequency_display()
        )

        worksheet.cell(
            row=row_num,
            column=5
        ).value = responsible_persons

        worksheet.cell(
            row=row_num,
            column=6
        ).value = reviewers

        worksheet.cell(
            row=row_num,
            column=7
        ).value = (
            submission.requirement.due_date.strftime('%d-%m-%Y')

            if submission.requirement.due_date
            else ''
        )

        worksheet.cell(
            row=row_num,
            column=8
        ).value = (
            submission.requirement.get_status_display()
        )

        worksheet.cell(
            row=row_num,
            column=9
        ).value = (
            submission.submitted_by.get_full_name()
        )

        worksheet.cell(
            row=row_num,
            column=10
        ).value = (

            submission.submitted_at.strftime(
                '%d-%m-%Y %H:%M'
            )

            if submission.submitted_at
            else ''
        )

        worksheet.cell(
            row=row_num,
            column=11
        ).value = (

            submission.reviewed_by.get_full_name()

            if submission.reviewed_by
            else ''
        )

        worksheet.cell(
            row=row_num,
            column=12
        ).value = (

            submission.reviewed_at.strftime(
                '%d-%m-%Y %H:%M'
            )

            if submission.reviewed_at
            else ''
        )

        row_num += 1

    # =================================================
    # AUTO COLUMN WIDTH
    # =================================================

    for column_cells in worksheet.columns:

        length = max(

            len(str(cell.value))
            if cell.value else 0

            for cell in column_cells
        )

        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 5

    # =================================================
    # RESPONSE
    # =================================================

    response = HttpResponse(

        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )

    response['Content-Disposition'] = (

        'attachment; filename=Compliance_Status_Report.xlsx'
    )

    workbook.save(response)

    return response



# =====================================================
# EXPORT COMPLIANCE STATUS REPORT PDF
# =====================================================

@login_required
def export_compliance_status_pdf(request):

    submissions = (

        ComplianceSubmission.objects

        .select_related(
            'requirement',
            'requirement__legal_act',
            'submitted_by',
            'reviewed_by'
        )

        .prefetch_related(
            'requirement__responsible_person',
            'requirement__reviewer'
        )

        .order_by('-created_at')
    )

    # =================================================
    # RESPONSE
    # =================================================

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename="Compliance_Status_Report.pdf"'
    )

    # =================================================
    # PDF DOCUMENT
    # =================================================

    document = SimpleDocTemplate(

        response,

        pagesize=landscape(A4),

        rightMargin=20,

        leftMargin=20,

        topMargin=20,

        bottomMargin=20
    )

    elements = []

    styles = getSampleStyleSheet()

    # =================================================
    # TITLE
    # =================================================

    title = Paragraph(

        '<b>Compliance Status Report</b>',

        styles['Title']
    )

    elements.append(title)

    elements.append(
        Spacer(1, 15)
    )

    # =================================================
    # TABLE DATA
    # =================================================

    data = [[

        'Code',

        'Compliance',

        'Legal Act',

        'Frequency',

        'Responsible',

        'Reviewer',

        'Due Date',

        'Status',

        'Submitted By',

        'Submitted At'
    ]]

    for submission in submissions:

        responsible_persons = ", ".join([

            user.get_full_name()

            for user in submission.requirement.responsible_person.all()
        ])

        reviewers = ", ".join([

            user.get_full_name()

            for user in submission.requirement.reviewer.all()
        ])

        data.append([

            submission.requirement.requirement_code,

            submission.requirement.title,

            submission.requirement.legal_act.short_name,

            submission.requirement.get_frequency_display(),

            responsible_persons,

            reviewers,

            submission.requirement.due_date.strftime(
                '%d-%m-%Y'
            ) if submission.requirement.due_date else '',

            submission.requirement.get_status_display(),

            submission.submitted_by.get_full_name(),

            submission.submitted_at.strftime(
                '%d-%m-%Y %H:%M'
            ) if submission.submitted_at else ''
        ])

    # =================================================
    # TABLE
    # =================================================

    table = Table(data)

    table.setStyle(TableStyle([

        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),

        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('FONTSIZE', (0, 0), (-1, -1), 8),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),

        ('GRID', (0, 0), (-1, -1), 1, colors.grey),

        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)

    # =================================================
    # BUILD PDF
    # =================================================

    document.build(elements)

    return response



# =====================================================
# OVERDUE COMPLIANCE DASHBOARD
# =====================================================

class OverdueComplianceDashboardView(
    LoginRequiredMixin,
    TemplateView
):

    template_name = (
        'legal_compliance/reports/overdue_dashboard.html'
    )

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        today = timezone.now().date()

        compliances = (

            ComplianceRequirement.objects

            .filter(
                is_active=True,
                due_date__lt=today
            )

            .exclude(
                status='COMPLETED'
            )

            .select_related(
                'legal_act'
            )

            .prefetch_related(
                'responsible_person',
                'reviewer'
            )

            .order_by('due_date')
        )

        # =============================================
        # CALCULATE DELAY DAYS
        # =============================================

        for compliance in compliances:

            compliance.delay_days = (
                today - compliance.due_date
            ).days

            # =========================================
            # ESCALATION LEVEL
            # =========================================

            if compliance.delay_days >= 30:

                compliance.escalation_level = (
                    'Management Escalation'
                )

            elif compliance.delay_days >= 15:

                compliance.escalation_level = (
                    'Function Head Escalation'
                )

            elif compliance.delay_days >= 7:

                compliance.escalation_level = (
                    'Reviewer Escalation'
                )

            else:

                compliance.escalation_level = (
                    'Responsible Person'
                )

        # =============================================
        # STATS
        # =============================================

        context['stats'] = {

            'total_overdue': compliances.count(),

            'critical_overdue': compliances.filter(
                criticality='HIGH'
            ).count(),

            'over_30_days': len([

                c for c in compliances

                if c.delay_days >= 30
            ]),

            'rejected': compliances.filter(
                status='REJECTED'
            ).count(),

            'pending_approval': compliances.filter(
                status='SUBMITTED'
            ).count(),
        }

        context['compliances'] = compliances

        return context