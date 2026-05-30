from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .models import EnvironmentalQuestion, MonthlyIndicatorAttachment, MonthlyIndicatorData


class EnvironmentalDataFetcher:
    """
    Dynamic data fetcher for auto-calculated environmental questions.
    """

    @classmethod
    def get_data_for_plant_year(cls, plant, year):
        result = {}
        auto_questions = EnvironmentalQuestion.objects.filter(
            is_active=True,
            source_type__in=["INCIDENT", "HAZARD", "INSPECTION"],
        )

        months = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        ]

        for question in auto_questions:
            month_data = {}
            for month_num, month_name in enumerate(months, start=1):
                month_data[month_name] = cls.calculate_question_value(
                    question,
                    plant,
                    month_num,
                    year,
                )
            result[question.question_text] = month_data

        return result

    @classmethod
    def calculate_question_value(cls, question, plant, month, year):
        if question.source_type == "INCIDENT":
            from apps.accidents.models import Incident

            queryset = Incident.objects.filter(
                plant=plant,
                incident_date__month=month,
                incident_date__year=year,
            )

            if question.filter_field == "incident_type" and question.filter_value:
                queryset = queryset.filter(incident_type_id=question.filter_value)
            elif question.filter_field == "status" and question.filter_value:
                queryset = queryset.filter(status=question.filter_value)
            elif question.filter_field == "plant" and question.filter_value:
                queryset = queryset.filter(plant_id=question.filter_value)

            if question.filter_field_2 and question.filter_value_2:
                if question.filter_field_2 == "incident_type":
                    queryset = queryset.filter(incident_type_id=question.filter_value_2)
                elif question.filter_field_2 == "status":
                    queryset = queryset.filter(status=question.filter_value_2)
                elif question.filter_field_2 == "plant":
                    queryset = queryset.filter(plant_id=question.filter_value_2)

            return queryset.count()

        if question.source_type == "HAZARD":
            from apps.hazards.models import Hazard

            queryset = Hazard.objects.filter(
                plant=plant,
                incident_datetime__year=year,
                incident_datetime__month=month,
            )

            if question.filter_field == "hazard_type" and question.filter_value:
                queryset = queryset.filter(hazard_type=question.filter_value)
            elif question.filter_field == "severity" and question.filter_value:
                queryset = queryset.filter(severity=question.filter_value)
            elif question.filter_field == "status" and question.filter_value:
                queryset = queryset.filter(status=question.filter_value)
            elif question.filter_field == "plant" and question.filter_value:
                queryset = queryset.filter(plant_id=question.filter_value)

            if question.filter_field_2 and question.filter_value_2:
                if question.filter_field_2 == "hazard_type":
                    queryset = queryset.filter(hazard_type=question.filter_value_2)
                elif question.filter_field_2 == "severity":
                    queryset = queryset.filter(severity=question.filter_value_2)
                elif question.filter_field_2 == "status":
                    queryset = queryset.filter(status=question.filter_value_2)
                elif question.filter_field_2 == "plant":
                    queryset = queryset.filter(plant_id=question.filter_value_2)

            return queryset.count()

        if question.source_type == "INSPECTION":
            from apps.inspections.models import InspectionSchedule

            queryset = InspectionSchedule.objects.filter(
                plants=plant,
                scheduled_date__month=month,
                scheduled_date__year=year,
            )

            if question.filter_field and question.filter_value:
                if question.filter_field == "template":
                    queryset = queryset.filter(template_id=question.filter_value)
                elif question.filter_field == "inspection_type":
                    queryset = queryset.filter(template__inspection_type=question.filter_value)
                elif question.filter_field == "status":
                    queryset = queryset.filter(status=question.filter_value)
                elif question.filter_field == "plant":
                    queryset = queryset.filter(plants__id=question.filter_value)
                elif question.filter_field == "assigned_to":
                    queryset = queryset.filter(assigned_to_id=question.filter_value)

            if question.filter_field_2 and question.filter_value_2:
                if question.filter_field_2 == "template":
                    queryset = queryset.filter(template_id=question.filter_value_2)
                elif question.filter_field_2 == "inspection_type":
                    queryset = queryset.filter(template__inspection_type=question.filter_value_2)
                elif question.filter_field_2 == "status":
                    queryset = queryset.filter(status=question.filter_value_2)
                elif question.filter_field_2 == "plant":
                    queryset = queryset.filter(plants__id=question.filter_value_2)
                elif question.filter_field_2 == "assigned_to":
                    queryset = queryset.filter(assigned_to_id=question.filter_value_2)

            return queryset.distinct().count()

        return 0


def get_financial_year_start_year(selected_fy=None):
    if selected_fy:
        try:
            return int(str(selected_fy).split("-")[0])
        except (TypeError, ValueError, AttributeError):
            pass

    today = datetime.now()
    return today.year - 1 if today.month < 4 else today.year


def get_financial_year_label(fy_start_year):
    return f"{fy_start_year}-{fy_start_year + 1}"


def get_financial_year_options(count=5, base_start_year=None):
    if base_start_year is None:
        base_start_year = get_financial_year_start_year()
    return [get_financial_year_label(base_start_year - offset) for offset in range(count)]


def get_financial_year_months(fy_start_year=None):
    fy_start_year = get_financial_year_start_year(fy_start_year)
    month_labels = dict(MonthlyIndicatorData.MONTH_CHOICES)
    month_codes = [
        "APR", "MAY", "JUN", "JUL", "AUG", "SEP",
        "OCT", "NOV", "DEC", "JAN", "FEB", "MAR",
    ]

    months = []
    for month_code in month_codes:
        month_name = month_labels.get(month_code, month_code.title())
        year = fy_start_year + 1 if month_code in ["JAN", "FEB", "MAR"] else fy_start_year
        months.append({
            "code": month_code,
            "label": f"{month_name} {year}",
            "month_name": month_name,
            "month_number": datetime.strptime(month_code, "%b").month,
            "year": year,
        })
    return months


def parse_decimal(value):
    if value in [None, "", "-"]:
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return None


def format_decimal(value):
    if value in [None, ""]:
        return "-"
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    normalized = value.normalize()
    return format(normalized, "f").rstrip("0").rstrip(".") or "0"


def get_environmental_questions():
    return list(
        EnvironmentalQuestion.objects.filter(is_active=True)
        .select_related("unit_category", "default_unit")
        .order_by("is_system", "order", "id")
    )


def build_environmental_report(plants, fy_start_year=None, selected_month=None, aggregate=False, include_attachments=True):
    plants = list(plants)
    questions = get_environmental_questions()
    months = get_financial_year_months(fy_start_year)

    if selected_month and selected_month != "all":
        months = [month for month in months if month["code"] == selected_month]

    manual_entries = MonthlyIndicatorData.objects.filter(
        plant__in=plants,
        indicator__in=questions,
    ).select_related("plant", "indicator", "unit")

    manual_dict = {}
    for entry in manual_entries:
        manual_dict.setdefault(entry.plant_id, {}).setdefault(entry.indicator_id, {})[entry.month] = entry.value

    attachments_dict = {}
    if include_attachments:
        attachments = MonthlyIndicatorAttachment.objects.filter(
            plant__in=plants,
            indicator__in=questions,
        ).select_related("plant", "indicator")
        for attachment in attachments:
            attachments_dict.setdefault(attachment.plant_id, {}).setdefault(attachment.indicator_id, {})[
                attachment.month
            ] = attachment

    def get_question_value(question, plant, month):
        if question.source_type == "MANUAL":
            return manual_dict.get(plant.id, {}).get(question.id, {}).get(month["code"])
        return EnvironmentalDataFetcher.calculate_question_value(
            question,
            plant,
            month["month_number"],
            month["year"],
        )

    if aggregate:
        questions_data = []
        for question in questions:
            month_data = {}
            annual_total = Decimal("0")
            has_values = False

            for month in months:
                month_total = Decimal("0")
                month_has_values = False

                for plant in plants:
                    numeric_value = parse_decimal(get_question_value(question, plant, month))
                    if numeric_value is not None:
                        month_total += numeric_value
                        month_has_values = True

                month_data[month["label"]] = {
                    "value": format_decimal(month_total) if month_has_values else "-",
                    "attachment": None,
                }

                if month_has_values:
                    annual_total += month_total
                    has_values = True

            questions_data.append({
                "question": question.question_text,
                "unit": question.default_unit.name if question.default_unit else "Count",
                "month_data": month_data,
                "annual": format_decimal(annual_total) if has_values else "-",
            })

        return {"months": months, "questions_data": questions_data}

    plants_data = []
    for plant in plants:
        question_rows = []
        for question in questions:
            month_data = {}
            annual_total = Decimal("0")
            has_values = False

            for month in months:
                raw_value = get_question_value(question, plant, month)
                numeric_value = parse_decimal(raw_value)
                month_data[month["label"]] = {
                    "value": raw_value if raw_value not in [None, ""] else "-",
                    "attachment": attachments_dict.get(plant.id, {}).get(question.id, {}).get(month["code"])
                    if include_attachments else None,
                }

                if numeric_value is not None:
                    annual_total += numeric_value
                    has_values = True

            question_rows.append({
                "question": question.question_text,
                "unit": question.default_unit.name if question.default_unit else "Count",
                "month_data": month_data,
                "annual": format_decimal(annual_total) if has_values else "-",
            })

        plants_data.append({"plant": plant, "questions_data": question_rows})

    return {"months": months, "plants_data": plants_data}
def generate_environmental_excel(report_data, title="Environmental Data"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Environmental Data"

    month_labels = [month["label"] for month in report_data["months"]]
    headers = ["Indicators"] + month_labels + ["Total"]

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    worksheet.cell(row=1, column=1, value=title)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=2, column=column_index, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_index, row in enumerate(report_data["questions_data"], start=3):
        label_cell = worksheet.cell(row=row_index, column=1, value=f"{row['question']} ({row['unit']})")
        label_cell.border = thin_border

        for month_index, month_label in enumerate(month_labels, start=2):
            value = row["month_data"].get(month_label, {}).get("value", "-")
            numeric_value = parse_decimal(value)
            cell = worksheet.cell(
                row=row_index,
                column=month_index,
                value=float(numeric_value) if numeric_value is not None else value,
            )
            cell.alignment = right_align if numeric_value is not None else center_align
            cell.border = thin_border

        total_value = row.get("annual", "-")
        numeric_total = parse_decimal(total_value)
        total_cell = worksheet.cell(
            row=row_index,
            column=len(headers),
            value=float(numeric_total) if numeric_total is not None else total_value,
        )
        total_cell.alignment = right_align if numeric_total is not None else center_align
        total_cell.border = thin_border

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max_length + 3

    worksheet.freeze_panes = "B3"
    return workbook
